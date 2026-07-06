"""
       ██╗███╗   ███╗ ██████╗
       ██║████╗ ████║██╔════╝
       ██║██╔████╔██║██║
  ██   ██║██║╚██╔╝██║██║    
  █████╔╝ ██║ ╚═╝ ██║╚██████╗
  ╚════╝  ╚═╝     ╚═╝ ╚═════╝
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import sqlite3
import os
from datetime import datetime
from PIL import Image
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import webbrowser
import time
import sys
import shutil # Para copias de seguridad
import logging # Para registro de actividad
import hashlib # Para hashing de contraseñas (básico, para ejemplo)

# --- Importaciones para Gráficos ---
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
import numpy as np

# --- Configuración de Logging ---
# Configura el logger para escribir en un archivo y en la consola
log_file = 'bim_manager_app.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)

# --- Configuración de CustomTkinter ---
ctk.set_appearance_mode("Dark")  # Modos: "System" (por defecto), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Temas: "blue" (por defecto), "green", "dark-blue"

# --- Configuración de la Base de Datos ---
DB_NAME = 'bim_projects.db'

def hash_password(password):
    """Hashea la contraseña usando SHA256."""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(stored_password, provided_password):
    """Verifica si la contraseña proporcionada coincide con la almacenada (hasheada)."""
    return stored_password == hashlib.sha256(provided_password.encode()).hexdigest()

def setup_database():
    """
    Configura la base de datos SQLite y crea las tablas 'projects', 'users' y 'project_updates_history'
    si no existen, o migra sus esquemas si es necesario.
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # Define el esquema deseado para la tabla 'projects'
    desired_projects_columns_definition = [
        "id INTEGER PRIMARY KEY AUTOINCREMENT",
        "codigo TEXT NOT NULL",
        "cadena TEXT NOT NULL",
        "pais TEXT NOT NULL",
        "ciudad TEXT NOT NULL",
        "ubicacion TEXT NOT NULL",
        "designacion_tienda TEXT NOT NULL",
        "nombre_completo TEXT",
        "image_model_version TEXT",
        "image_model_date TEXT",
        "constructive_model_version TEXT",
        "constructive_model_date TEXT",
        "mobiliario_incorporado INTEGER",
        "mobiliario_version_propuesta TEXT",
        "mobiliario_fecha_propuesta TEXT",
        "mobiliario_actualizacion_date TEXT",
        "mobiliario_incorporation_date TEXT",
        "puertas_incorporadas INTEGER",
        "puertas_version_propuesta TEXT",
        "puertas_fecha_propuesta TEXT",
        "puertas_actualizacion_date TEXT",
        "puertas_incorporation_date TEXT",
        "fecha_inicio_obra TEXT",
        "fecha_apertura TEXT",
        "creation_date TEXT",
        "revision_version INTEGER DEFAULT 0" # Nueva columna para la versión de revisión
    ]
    
    # Nombres de las columnas en el esquema deseado para 'projects'
    desired_projects_column_names = [col.split(' ')[0] for col in desired_projects_columns_definition]

    # Define el esquema deseado para la tabla 'users'
    desired_users_columns_definition = [
        "id INTEGER PRIMARY KEY AUTOINCREMENT",
        "username TEXT UNIQUE NOT NULL",
        "password TEXT NOT NULL", # Almacenar el hash de la contraseña
        "role TEXT NOT NULL DEFAULT 'Visitor'" # 'Admin', 'Editor', 'Visitor'
    ]
    desired_users_column_names = [col.split(' ')[0] for col in desired_users_columns_definition]

    # Define el esquema deseado para la tabla 'project_updates_history'
    desired_history_columns_definition = [
        "id INTEGER PRIMARY KEY AUTOINCREMENT",
        "project_id INTEGER NOT NULL",
        "user_id INTEGER", # Nuevo: Link a la tabla users
        "peticionario_name TEXT", # Nuevo: Nombre del peticionario (no el usuario logueado)
        "revision_version_at_update INTEGER",
        "update_date TEXT",
        "changed_field TEXT",   # Nuevo: campo que se cambió
        "old_value TEXT",       # Nuevo: valor anterior
        "new_value TEXT",       # Nuevo: nuevo valor
        "FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE",
        "FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL" # Si el usuario es eliminado, establecer a NULL
    ]
    
    # Nombres de las columnas en el esquema deseado para 'project_updates_history'
    desired_history_column_names = [col.split(' ')[0] for col in desired_history_columns_definition if not col.startswith("FOREIGN KEY")]


    # --- Paso 1: Limpiar cualquier tabla temporal de migraciones fallidas anteriores ---
    try:
        cursor.execute("DROP TABLE IF EXISTS projects_temp")
        cursor.execute("DROP TABLE IF EXISTS project_updates_history_temp") 
        cursor.execute("DROP TABLE IF EXISTS users_temp") # También limpiar tabla temporal de usuarios
        conn.commit()
    except sqlite3.Error as e:
        logging.warning(f"Advertencia: No se pudieron eliminar tablas temporales (puede que no existieran): {e}")

    # --- Paso 2: Obtener el esquema actual de la tabla 'projects' y migrar si es necesario ---
    cursor.execute("PRAGMA table_info(projects)")
    existing_projects_columns_info = cursor.fetchall()
    existing_projects_column_names = [col[1] for col in existing_projects_columns_info]

    needs_projects_migration = False
    if not existing_projects_column_names:
        logging.info("La tabla 'projects' no existe. Creándola con el esquema más reciente.")
        create_table_sql = f"CREATE TABLE projects ({', '.join(desired_projects_columns_definition)})"
        cursor.execute(create_table_sql)
        conn.commit()
    else:
        for col_name in desired_projects_column_names:
            if col_name not in existing_projects_column_names:
                needs_projects_migration = True
                break
        
        if needs_projects_migration:
            logging.info("Detectada tabla 'projects' con esquema antiguo o incompleto. Iniciando migración...")
            try:
                create_temp_table_sql = f"CREATE TABLE projects_temp ({', '.join(desired_projects_columns_definition)})"
                cursor.execute(create_temp_table_sql)
                conn.commit()

                common_columns = [col for col in existing_projects_column_names if col in desired_projects_column_names]
                
                if common_columns:
                    cols_str = ", ".join(common_columns)
                    insert_placeholders = ", ".join(["?" for _ in common_columns])
                    
                    cursor.execute(f"SELECT {cols_str} FROM projects")
                    data_to_migrate = cursor.fetchall()

                    for row in data_to_migrate:
                        if len(row) == len(common_columns):
                            cursor.execute(f"INSERT INTO projects_temp ({cols_str}) VALUES ({insert_placeholders})", row)
                        else:
                            logging.warning(f"Saltando fila por discrepancia de columnas durante la migración de projects: {row}")
                    conn.commit()
                    logging.info("Datos migrados a projects_temp.")

                cursor.execute("DROP TABLE projects")
                conn.commit()
                logging.info("Tabla 'projects' antigua eliminada.")

                cursor.execute("ALTER TABLE projects_temp RENAME TO projects")
                conn.commit()
                logging.info("Tabla 'projects_temp' renombrada a 'projects'. Migración completada para 'projects'.")

            except sqlite3.Error as e:
                logging.error(f"Error crítico durante la migración de la base de datos 'projects': {e}")
                try:
                    cursor.execute("DROP TABLE IF EXISTS projects_temp")
                    conn.commit()
                    logging.info("Limpieza de tabla temporal 'projects_temp' realizada tras error de migración.")
                except sqlite3.Error as cleanup_e:
                    logging.error(f"Error al intentar limpiar 'projects_temp': {cleanup_e}")
                raise 
        else:
            logging.info("Tabla 'projects' ya tiene el esquema más reciente. No se necesita migración.")

    # --- Paso 3: Obtener el esquema actual de la tabla 'users' y migrar si es necesario ---
    cursor.execute("PRAGMA table_info(users)")
    existing_users_columns_info = cursor.fetchall()
    existing_users_column_names = [col[1] for col in existing_users_columns_info]

    needs_users_migration = False
    if not existing_users_column_names:
        logging.info("La tabla 'users' no existe. Creándola.")
        create_users_table_sql = f"CREATE TABLE users ({', '.join(desired_users_columns_definition)})"
        cursor.execute(create_users_table_sql)
        conn.commit()
    else:
        for col_name in desired_users_column_names:
            if col_name not in existing_users_column_names:
                needs_users_migration = True
                break
        
        if needs_users_migration:
            logging.info("Detectada tabla 'users' con esquema antiguo o incompleto. Iniciando migración...")
            try:
                create_temp_users_table_sql = f"CREATE TABLE users_temp ({', '.join(desired_users_columns_definition)})"
                cursor.execute(create_temp_users_table_sql)
                conn.commit()

                common_users_columns = [col for col in existing_users_column_names if col in desired_users_column_names]
                
                if common_users_columns:
                    cols_str = ", ".join(common_users_columns)
                    insert_placeholders = ", ".join(["?" for _ in common_users_columns])
                    
                    cursor.execute(f"SELECT {cols_str} FROM users")
                    users_data_to_migrate = cursor.fetchall()

                    for row in users_data_to_migrate:
                        if len(row) == len(common_users_columns):
                            cursor.execute(f"INSERT INTO users_temp ({cols_str}) VALUES ({insert_placeholders})", row)
                        else:
                            logging.warning(f"Saltando fila por discrepancia de columnas durante la migración de users: {row}")
                    conn.commit()
                    logging.info("Datos de usuarios migrados a users_temp.")

                cursor.execute("DROP TABLE users")
                conn.commit()
                logging.info("Tabla 'users' antigua eliminada.")

                cursor.execute("ALTER TABLE users_temp RENAME TO users")
                conn.commit()
                logging.info("Tabla 'users_temp' renombrada a 'users'. Migración completada para 'users'.")

            except sqlite3.Error as e:
                logging.error(f"Error crítico durante la migración de la base de datos 'users': {e}")
                try:
                    cursor.execute("DROP TABLE IF EXISTS users_temp")
                    conn.commit()
                    logging.info("Limpieza de tabla temporal 'users_temp' realizada tras error de migración.")
                except sqlite3.Error as cleanup_e:
                    logging.error(f"Error al intentar limpiar 'users_temp': {cleanup_e}")
                raise 
        else:
            logging.info("Tabla 'users' ya tiene el esquema más reciente. No se necesita migración.")

    # --- Paso 4: Obtener el esquema actual de la tabla 'project_updates_history' y migrar si es necesario ---
    cursor.execute("PRAGMA table_info(project_updates_history)")
    existing_history_columns_info = cursor.fetchall()
    existing_history_column_names = [col[1] for col in existing_history_columns_info]

    needs_history_migration = False
    if not existing_history_column_names:
        logging.info("La tabla 'project_updates_history' no existe. Creándola.")
        create_history_table_sql = f"CREATE TABLE project_updates_history ({', '.join(desired_history_columns_definition)})"
        cursor.execute(create_history_table_sql)
        conn.commit()
    else:
        for col_name in desired_history_column_names:
            if col_name not in existing_history_column_names:
                needs_history_migration = True
                break
        
        if needs_history_migration:
            logging.info("Detectada tabla 'project_updates_history' con esquema antiguo o incompleto. Iniciando migración...")
            try:
                create_temp_history_table_sql = f"CREATE TABLE project_updates_history_temp ({', '.join(desired_history_columns_definition)})"
                cursor.execute(create_temp_history_table_sql)
                conn.commit()

                common_history_columns = [col for col in existing_history_column_names if col in desired_history_column_names]
                
                if common_history_columns:
                    cols_str = ", ".join(common_history_columns)
                    insert_placeholders = ", ".join(["?" for _ in common_columns])
                    
                    cursor.execute(f"SELECT {cols_str} FROM project_updates_history")
                    history_data_to_migrate = cursor.fetchall()

                    for row in history_data_to_migrate:
                        if len(row) == len(common_history_columns):
                            cursor.execute(f"INSERT INTO project_updates_history_temp ({cols_str}) VALUES ({insert_placeholders})", row)
                        else:
                            logging.warning(f"Saltando fila por discrepancia de columnas durante la migración de history: {row}")
                    conn.commit()
                    logging.info("Datos migrados a project_updates_history_temp.")

                cursor.execute("DROP TABLE project_updates_history")
                conn.commit()
                logging.info("Tabla 'project_updates_history' antigua eliminada.")

                cursor.execute("ALTER TABLE project_updates_history_temp RENAME TO project_updates_history")
                conn.commit()
                logging.info("Tabla 'project_updates_history_temp' renombrada a 'project_updates_history'. Migración completada para 'project_updates_history'.")

            except sqlite3.Error as e:
                logging.error(f"Error crítico durante la migración de la base de datos 'project_updates_history': {e}")
                try:
                    cursor.execute("DROP TABLE IF EXISTS project_updates_history_temp")
                    conn.commit()
                    logging.info("Limpieza de tabla temporal 'project_updates_history_temp' realizada tras error de migración.")
                except sqlite3.Error as cleanup_e:
                    logging.error(f"Error al intentar limpiar 'project_updates_history_temp': {cleanup_e}")
                raise 
        else:
            logging.info("Tabla 'project_updates_history' ya tiene el esquema más reciente. No se necesita migración.")

    # --- Crear usuario administrador por defecto si no existen usuarios ---
    cursor.execute("SELECT COUNT(*) FROM users")
    user_count = cursor.fetchone()[0]
    if user_count == 0:
        logging.info("No se encontraron usuarios. Creando usuario administrador por defecto (admin/admin123).")
        try:
            cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                           ('admin', hash_password('admin123'), 'Admin'))
            conn.commit()
            logging.info("Usuario 'admin' creado exitosamente.")
        except sqlite3.Error as e:
            logging.error(f"Error al crear el usuario administrador por defecto: {e}")

    conn.close()
    logging.info(f"Base de datos '{DB_NAME}' configurada correctamente.")


def is_valid_date(date_string):
    """
    Verifica si una cadena de texto es una fecha válida en formato YYYY-MM-DD.
    """
    if not date_string: # Permitir fechas vacías
        return True
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def get_all_projects():
    """
    Recupera todos los proyectos de la base de datos.
    Retorna una lista de tuplas (id, codigo, ciudad, nombre_completo).
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT id, codigo, ciudad, nombre_completo FROM projects ORDER BY codigo ASC")
    projects = cursor.fetchall()
    conn.close()
    return projects

def get_project_details(project_id):
    """
    Recupera todos los detalles de un proyecto por su ID.
    Retorna un diccionario con los detalles del proyecto.
    """
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row # Use row_factory for dictionary-like access
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM projects WHERE id = ?", (project_id,))
    project_data = cursor.fetchone()
    conn.close()

    if project_data:
        return dict(project_data) # Convert Row object to dictionary
    return None

def get_project_update_history(project_id):
    """
    Recupera el historial de actualizaciones de un proyecto específico, incluyendo los detalles del cambio.
    Retorna una lista de tuplas (username, peticionario_name, revision_version_at_update, update_date, changed_field, old_value, new_value).
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            u.username, 
            puh.peticionario_name,
            puh.revision_version_at_update, 
            puh.update_date,
            puh.changed_field,
            puh.old_value,
            puh.new_value
        FROM project_updates_history AS puh
        LEFT JOIN users AS u ON puh.user_id = u.id
        WHERE puh.project_id = ?
        ORDER BY puh.update_date DESC, puh.revision_version_at_update DESC
    """, (project_id,))
    history = cursor.fetchall()
    conn.close()
    return history

def get_all_users():
    """
    Recupera todos los usuarios de la base de datos.
    Retorna una lista de tuplas (id, username, role).
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, role FROM users ORDER BY username ASC")
    users = cursor.fetchall()
    conn.close()
    return users

def add_user_to_db(username, password, role):
    """Añade un nuevo usuario a la base de datos."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                       (username, hash_password(password), role))
        conn.commit()
        logging.info(f"Usuario '{username}' con rol '{role}' añadido exitosamente.")
        return True
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "El nombre de usuario ya existe.")
        logging.warning(f"Intento de añadir usuario '{username}' fallido: el nombre de usuario ya existe.")
        return False
    except sqlite3.Error as e:
        logging.error(f"Error al añadir usuario '{username}': {e}")
        messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al añadir el usuario: {e}")
        return False
    finally:
        conn.close()

def update_user_in_db(user_id, username, password, role):
    """Actualiza un usuario existente en la base de datos."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        if password: # Solo actualiza la contraseña si se proporciona una nueva
            cursor.execute("UPDATE users SET username = ?, password = ?, role = ? WHERE id = ?",
                           (username, hash_password(password), role, user_id))
        else:
            cursor.execute("UPDATE users SET username = ?, role = ? WHERE id = ?",
                           (username, role, user_id))
        conn.commit()
        logging.info(f"Usuario '{username}' (ID: {user_id}) actualizado exitosamente con rol '{role}'.")
        return True
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "El nombre de usuario ya existe.")
        logging.warning(f"Intento de actualizar usuario (ID: {user_id}) a '{username}' fallido: el nombre de usuario ya existe.")
        return False
    except sqlite3.Error as e:
        logging.error(f"Error al actualizar usuario (ID: {user_id}): {e}")
        messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al actualizar el usuario: {e}")
        return False
    finally:
        conn.close()

def delete_user_from_db(user_id):
    """Elimina un usuario de la base de datos."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        logging.info(f"Usuario (ID: {user_id}) eliminado exitosamente.")
        return True
    except sqlite3.Error as e:
        logging.error(f"Error al eliminar usuario (ID: {user_id}): {e}")
        messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al eliminar el usuario: {e}")
        return False
    finally:
        conn.close()


def search_projects_in_db(global_search_term="", advanced_field=None, advanced_term=None):
    """
    Busca proyectos en la base de datos por un término global o un campo específico.
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    query = "SELECT id, nombre_completo, image_model_version, mobiliario_version_propuesta, puertas_version_propuesta, fecha_inicio_obra, fecha_apertura FROM projects WHERE 1=1"
    params = []

    if advanced_field and advanced_term:
        if advanced_field == "peticionario_name":
            # Subquery to find project_ids based on peticionario_name
            cursor.execute("SELECT DISTINCT project_id FROM project_updates_history WHERE peticionario_name LIKE ?", (f"%{advanced_term}%",))
            project_ids = [row[0] for row in cursor.fetchall()]
            if not project_ids:
                return [] # No projects found for this peticionario
            # Use IN clause for the found project IDs
            query += f" AND id IN ({','.join('?' for _ in project_ids)})"
            params.extend(project_ids)
        else:
            query += f" AND {advanced_field} LIKE ?"
            params.append(f"%{advanced_term}%")

    elif global_search_term:
        search_pattern = f"%{global_search_term}%"
        search_fields = [
            "codigo", "cadena", "pais", "ciudad", "ubicacion", "designacion_tienda",
            "nombre_completo", "image_model_version", "constructive_model_version",
            "mobiliario_version_propuesta", "puertas_version_propuesta",
            "fecha_inicio_obra", "fecha_apertura", "creation_date"
        ]
        or_clauses = [f"{field} LIKE ?" for field in search_fields]
        query += " AND (" + " OR ".join(or_clauses) + ")"
        params.extend([search_pattern] * len(search_fields))
    
    query += " ORDER BY codigo ASC"

    cursor.execute(query, tuple(params))
    results = cursor.fetchall()
    conn.close()
    return results


class UpdateProjectDialog(ctk.CTkToplevel):
    """
    Ventana de diálogo para actualizar múltiples datos de un proyecto.
    """
    # Mapping of internal field names to display names for history
    FIELD_DISPLAY_NAMES = {
        "image_model_version": "Modelo Imagen Versión",
        "image_model_date": "Fecha MI",
        "constructive_model_version": "Modelo Constructivo Versión",
        "constructive_model_date": "Fecha MC",
        "mobiliario_incorporado": "Mobiliario Incorporado",
        "mobiliario_version_propuesta": "Mobiliario Versión Propuesta",
        "mobiliario_fecha_propuesta": "Mobiliario Fecha Propuesta",
        "mobiliario_actualizacion_date": "Mobiliario Actualización",
        "mobiliario_incorporation_date": "Mobiliario 1ª Incorporación",
        "puertas_incorporadas": "Puertas Incorporadas",
        "puertas_version_propuesta": "Puertas Versión Propuesta",
        "puertas_fecha_propuesta": "Puertas Fecha Propuesta",
        "puertas_actualizacion_date": "Puertas Actualización",
        "puertas_incorporation_date": "Puertas 1ª Incorporación",
        "fecha_inicio_obra": "Fecha Inicio Obra",
        "fecha_apertura": "Fecha Apertura",
        "revision_version": "Versión Revisión",
        "codigo": "Código",
        "cadena": "Cadena",
        "pais": "País",
        "ciudad": "Ciudad",
        "ubicacion": "Ubicación",
        "designacion_tienda": "Designación Tienda",
        "nombre_completo": "Nombre Completo",
        "creation_date": "Fecha de Alta"
    }

    def __init__(self, parent, project_data, current_user_id, callback):
        super().__init__(parent)
        self.title("Actualizar Datos del Proyecto")
        self.geometry("700x850") # Aumentado el tamaño de la ventana
        self.grab_set() # Hace que la ventana sea modal
        self.resizable(False, False)

        # Intenta establecer el icono de la ventana (para Windows)
        try:
            # CORRECCIÓN: Usar una ruta absoluta o relativa robusta
            icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
            if os.path.exists(icon_path):
                self.wm_iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"Advertencia: No se pudo cargar 'icon.ico' para la ventana de actualización: {e}")

        self.project_data = project_data # Datos originales del proyecto
        self.current_user_id = current_user_id # ID del usuario que realiza la actualización
        self.callback = callback # Función a llamar después de la actualización

        # Configuración de grid para la ventana principal del diálogo
        self.grid_rowconfigure(1, weight=1) # Fila para el contenido scrollable
        self.grid_columnconfigure(0, weight=1)

        # Título del diálogo
        ctk.CTkLabel(self, text=f"Actualizar Proyecto: {project_data.get('nombre_completo', 'N/A')}", 
                     font=ctk.CTkFont(size=18, weight="bold")).grid(row=0, column=0, pady=15, padx=20, sticky="ew")

        # Frame scrollable para los campos de entrada
        self.scrollable_content_frame = ctk.CTkScrollableFrame(self, corner_radius=10)
        self.scrollable_content_frame.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        self.scrollable_content_frame.grid_columnconfigure(1, weight=1) # Columna para las entradas

        self.entries = {} # Diccionario para almacenar referencias a CTkEntry

        # --- Campo Peticionario (ahora editable) ---
        row_idx = 0
        ctk.CTkLabel(self.scrollable_content_frame, text="Peticionario:", font=ctk.CTkFont(weight="bold")).grid(row=row_idx, column=0, sticky="w", padx=10, pady=5)
        self.entry_peticionario = ctk.CTkEntry(self.scrollable_content_frame, width=250, placeholder_text="Nombre del compañero", font=ctk.CTkFont(weight="bold"))
        self.entry_peticionario.grid(row=row_idx, column=1, sticky="ew", padx=10, pady=5)
        # No pre-rellenar con el nombre de usuario, permitir entrada manual
        
        row_idx += 1
        ctk.CTkLabel(self.scrollable_content_frame, text="").grid(row=row_idx, column=0, columnspan=3, pady=5) # Separador visual
        row_idx += 1


        # Mostrar Versión Revisión (no editable)
        ctk.CTkLabel(self.scrollable_content_frame, text="Versión Revisión Actual:", font=ctk.CTkFont(weight="bold")).grid(row=row_idx, column=0, sticky="w", padx=10, pady=5)
        self.revision_version_label = ctk.CTkLabel(self.scrollable_content_frame, text=str(project_data.get('revision_version', 0)), font=ctk.CTkFont(weight="bold"))
        self.revision_version_label.grid(row=row_idx, column=1, sticky="w", padx=10, pady=5)
        row_idx += 1
        ctk.CTkLabel(self.scrollable_content_frame, text="").grid(row=row_idx, column=0, columnspan=3, pady=5) # Separador visual
        row_idx += 1

        # --- Campos para actualización manual (sin Revit) ---
        manual_fields = [
            ("Modelo Imagen Versión", "image_model_version", "image_model_date"),
            ("Modelo Constructivo Versión", "constructive_model_version", "constructive_model_date"),
            ("Mobiliario 1ª Inc.", None, "mobiliario_incorporation_date"), 
            ("Puertas 1ª Inc.", None, "puertas_incorporation_date"), 
            ("Fecha Inicio Obra", None, "fecha_inicio_obra"),
            ("Fecha Apertura", None, "fecha_apertura"),
        ]

        for label_text, version_key, date_key in manual_fields:
            ctk.CTkLabel(self.scrollable_content_frame, text=f"{label_text}:", font=ctk.CTkFont(weight="bold")).grid(row=row_idx, column=0, sticky="w", padx=10, pady=5)
            
            if version_key: 
                entry_version = ctk.CTkEntry(self.scrollable_content_frame, width=250, font=ctk.CTkFont(weight="bold")) # Ancho aumentado
                entry_version.insert(0, str(project_data.get(version_key, '')))
                entry_version.grid(row=row_idx, column=1, sticky="ew", padx=10, pady=5)
                self.entries[version_key] = entry_version
            else: 
                ctk.CTkLabel(self.scrollable_content_frame, text="", font=ctk.CTkFont(weight="bold")).grid(row=row_idx, column=1, sticky="w", padx=10, pady=5) 

            row_idx += 1
            if date_key: 
                ctk.CTkLabel(self.scrollable_content_frame, text=f"Fecha {label_text.split(' Versión')[0]} (YYYY-MM-DD):", font=ctk.CTkFont(weight="bold")).grid(row=row_idx, column=0, sticky="w", padx=10, pady=5)
                entry_date = ctk.CTkEntry(self.scrollable_content_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold")) # Ancho aumentado
                entry_date.insert(0, str(project_data.get(date_key, '')))
                entry_date.grid(row=row_idx, column=1, sticky="ew", padx=10, pady=5)
                self.entries[date_key] = entry_date
                
                var_update_date = ctk.BooleanVar()
                check_update_date = ctk.CTkCheckBox(self.scrollable_content_frame, text="Fecha Hoy", variable=var_update_date,
                                                    command=lambda e=entry_date, v=var_update_date: self._set_date_to_today(e, v), font=ctk.CTkFont(weight="bold"))
                check_update_date.grid(row=row_idx, column=2, sticky="w", padx=5)
                self.entries[f"var_update_{date_key}"] = var_update_date 
            row_idx += 1
            ctk.CTkLabel(self.scrollable_content_frame, text="").grid(row=row_idx, column=0, columnspan=3, pady=5) # Separador visual
            row_idx += 1


        # Frame para los botones (fuera del scrollable_content_frame)
        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=2, column=0, pady=20, sticky="ew", padx=20)
        button_frame.grid_columnconfigure(0, weight=1) 
        button_frame.grid_columnconfigure(1, weight=1) 

        btn_accept = ctk.CTkButton(button_frame, text="Aceptar", command=self._save_changes, fg_color="green", font=ctk.CTkFont(weight="bold"))
        btn_accept.grid(row=0, column=0, padx=10, sticky="e") 
        
        btn_cancel = ctk.CTkButton(button_frame, text="Cancelar", command=self.destroy, fg_color="red", font=ctk.CTkFont(weight="bold"))
        btn_cancel.grid(row=0, column=1, padx=10, sticky="w") 


    def _set_date_to_today(self, entry_widget, var_update_date):
        """Establece la fecha del campo de entrada a la fecha actual si el checkbox está marcado."""
        if var_update_date.get():
            entry_widget.delete(0, ctk.END)
            entry_widget.insert(0, datetime.now().strftime('%Y-%m-%d'))
        else:
            entry_widget.delete(0, ctk.END) # Limpiar si desmarca

    def _save_changes(self):
        """Guarda los cambios en la base de datos y cierra el diálogo."""
        updated_data = {}
        current_date_str = datetime.now().strftime('%Y-%m-%d')
        peticionario_name = self.entry_peticionario.get().strip()

        if not peticionario_name:
            messagebox.showerror("Error de Entrada", "El nombre del Peticionario no puede estar vacío.")
            return

        # Obtener los datos más recientes del proyecto para comparar y registrar cambios
        latest_project_data = get_project_details(self.project_data['id'])
        
        # Recopilar datos de los campos manuales
        for key, entry_widget in self.entries.items():
            if not key.startswith("var_update_") and key != 'peticionario_name':
                new_value = entry_widget.get().strip()
                updated_data[key] = new_value
                if "date" in key and new_value and not is_valid_date(new_value):
                    messagebox.showerror("Error de Fecha", f"La fecha '{new_value}' para {self.FIELD_DISPLAY_NAMES.get(key, key)} no tiene un formato válido (YYYY-MM-DD).")
                    return
        
        # --- Lógica de auto-actualización para Mobiliario y Puertas ---
        # Mobiliario
        current_mobiliario_version = latest_project_data.get('mobiliario_version_propuesta', '') # Use empty string for consistency
        new_mobiliario_version_input = updated_data.get('mobiliario_version_propuesta', '')

        if new_mobiliario_version_input != current_mobiliario_version and new_mobiliario_version_input not in ('', 'N/A'):
            try:
                next_mobiliario_version = str(int(current_mobiliario_version) + 1) if current_mobiliario_version.isdigit() else '1'
            except ValueError:
                next_mobiliario_version = '1' # Fallback if current version is not a number
            updated_data['mobiliario_version_propuesta'] = next_mobiliario_version
            updated_data['mobiliario_actualizacion_date'] = current_date_str
        else:
            # Si no hay cambio en la versión propuesta manual, mantener los valores existentes
            updated_data['mobiliario_version_propuesta'] = latest_project_data.get('mobiliario_version_propuesta', '')
            updated_data['mobiliario_actualizacion_date'] = latest_project_data.get('mobiliario_actualizacion_date', '')


        # Puertas
        current_puertas_version = latest_project_data.get('puertas_version_propuesta', '')
        new_puertas_version_input = updated_data.get('puertas_version_propuesta', '')

        if new_puertas_version_input != current_puertas_version and new_puertas_version_input not in ('', 'N/A'):
            try:
                next_puertas_version = str(int(current_puertas_version) + 1) if current_puertas_version.isdigit() else '1'
            except ValueError:
                next_puertas_version = '1'
            updated_data['puertas_version_propuesta'] = next_puertas_version
            updated_data['puertas_actualizacion_date'] = current_date_str
        else:
            updated_data['puertas_version_propuesta'] = latest_project_data.get('puertas_version_propuesta', '')
            updated_data['puertas_actualizacion_date'] = latest_project_data.get('puertas_actualizacion_date', '')


        # Incrementar Versión Revisión
        current_revision_version = latest_project_data.get('revision_version', 0)
        new_revision_version = current_revision_version + 1
        updated_data['revision_version'] = new_revision_version


        conn = None
        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            
            set_clauses = []
            values = []
            changes_made = [] # Para registrar los cambios detallados

            for key, new_value in updated_data.items():
                old_value = latest_project_data.get(key)
                # Convertir None a cadena vacía para comparación consistente
                old_value_str = str(old_value) if old_value is not None else ''
                new_value_str = str(new_value) if new_value is not None else ''

                if old_value_str != new_value_str:
                    set_clauses.append(f"{key} = ?")
                    values.append(new_value)
                    changes_made.append({
                        "field": key,
                        "old_value": old_value_str,
                        "new_value": new_value_str
                    })
            
            if not set_clauses:
                messagebox.showinfo("Información", "No hay cambios para guardar.")
                self.destroy()
                return

            values.append(self.project_data['id'])
            
            query = f"UPDATE projects SET {', '.join(set_clauses)} WHERE id = ?"
            cursor.execute(query, tuple(values))

            # Insertar registro(s) en project_updates_history para cada cambio individual
            for change in changes_made:
                display_field_name = self.FIELD_DISPLAY_NAMES.get(change['field'], change['field'].replace('_', ' ').title())
                cursor.execute('''
                    INSERT INTO project_updates_history (project_id, user_id, peticionario_name, revision_version_at_update, update_date, changed_field, old_value, new_value)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (self.project_data['id'], self.current_user_id, peticionario_name, new_revision_version, current_date_str, 
                      display_field_name, change['old_value'], change['new_value']))
            
            conn.commit()
            logging.info(f"Proyecto {self.project_data['id']} actualizado por usuario ID {self.current_user_id}. Peticionario: {peticionario_name}. Nueva versión: {new_revision_version}. Cambios: {changes_made}")
            messagebox.showinfo("Éxito", "Datos del proyecto actualizados correctamente.")
            self.callback(self.project_data['id']) # Refrescar la vista principal
            self.destroy()
        except sqlite3.Error as e:
            logging.error(f"Error de base de datos al actualizar proyecto {self.project_data['id']}: {e}")
            messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al actualizar los datos: {e}")
        finally:
            if conn:
                conn.close()

class WelcomeScreen(ctk.CTkToplevel):
    """
    Ventana de bienvenida inicial.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Bienvenido")
        self.geometry("500x400") # Increased height for better logo display
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.grab_set()

        # Center the window
        self.after(100, self.center_window)

        self.login_requested = False

        # Main frame with a blue border
        main_frame = ctk.CTkFrame(self, corner_radius=15, border_width=3, border_color="#2D68C4")
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        # Load and display logo
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "Jose.png")
            if os.path.exists(logo_path):
                # Corrected size to maintain aspect ratio for a 224x224 image
                self.logo_image = ctk.CTkImage(Image.open(logo_path), size=(200, 200)) 
                logo_label = ctk.CTkLabel(main_frame, image=self.logo_image, text="")
                logo_label.pack(pady=(20, 15))
            else:
                logging.warning("No se encontró el archivo Jose.png. Mostrando texto en su lugar.")
                ctk.CTkLabel(main_frame, text="Jose", font=ctk.CTkFont(size=30, weight="bold")).pack(pady=(30, 15))

        except Exception as e:
            logging.error(f"Error al cargar el logo en WelcomeScreen: {e}")
            ctk.CTkLabel(main_frame, text="Jose", font=ctk.CTkFont(size=30, weight="bold")).pack(pady=(30, 15))

        ctk.CTkLabel(main_frame, text="BIM Manager", font=ctk.CTkFont(size=28, weight="bold")).pack(pady=5)
        
        access_button = ctk.CTkButton(main_frame, text="Acceder", command=self.request_login, height=40, font=ctk.CTkFont(weight="bold"))
        access_button.pack(pady=(20, 30), padx=50, fill="x")

        self.focus_set()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f'+{x}+{y}')

    def request_login(self):
        self.login_requested = True
        self.destroy()

    def on_closing(self):
        self.login_requested = False
        self.destroy()
        sys.exit()

class LoginScreen(ctk.CTkToplevel):
    """
    Ventana de inicio de sesión para la aplicación.
    """
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Acceso")
        self.geometry("400x280")
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.grab_set() # Make the window modal

        self.login_successful = False
        self.user_id = None
        self.user_role = None

        main_frame = ctk.CTkFrame(self)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        ctk.CTkLabel(main_frame, text="Acceso a BIM Manager", font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(10, 20))
        
        ctk.CTkLabel(main_frame, text="Nombre de Usuario:").pack(padx=20, pady=(10,5), anchor="w")
        self.username_entry = ctk.CTkEntry(main_frame)
        self.username_entry.pack(fill="x", padx=20)

        ctk.CTkLabel(main_frame, text="Contraseña:").pack(padx=20, pady=(10,5), anchor="w")
        self.password_entry = ctk.CTkEntry(main_frame, show="*")
        self.password_entry.pack(fill="x", padx=20)
        self.password_entry.bind("<Return>", self.check_credentials)

        self.login_button = ctk.CTkButton(main_frame, text="Entrar", command=self.check_credentials)
        self.login_button.pack(pady=20, padx=20)
        
        self.focus_set() # Set focus to this window
        self.username_entry.focus() # Set focus to the username entry

    def check_credentials(self, event=None):
        """
        Verifica las credenciales introducidas contra la base de datos.
        """
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()

        if not username or not password:
            messagebox.showerror("Error de Credenciales", "Por favor, introduzca nombre de usuario y contraseña.")
            return

        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT id, password, role FROM users WHERE username = ?", (username,))
        user_data = cursor.fetchone()
        conn.close()

        if user_data and verify_password(user_data[1], password):
            self.login_successful = True
            self.user_id = user_data[0]
            self.user_role = user_data[2]
            logging.info(f"Inicio de sesión exitoso para el usuario '{username}' con rol '{self.user_role}'.")
            self.destroy()
        else:
            logging.warning(f"Intento de inicio de sesión fallido para el usuario '{username}'.")
            messagebox.showerror("Error de Credenciales", "Nombre de usuario o contraseña incorrectos.")
            self.password_entry.delete(0, "end") # Clear the password entry
            self.username_entry.focus() # Re-focus for retry

    def on_closing(self):
        """
        Maneja el evento de cierre de la ventana de login.
        """
        logging.info("Ventana de inicio de sesión cerrada. Saliendo de la aplicación.")
        self.destroy()
        sys.exit() # Exit the application if login window is closed

class AdvancedSearchDialog(ctk.CTkToplevel):
    """Dialog for advanced project search."""
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.title("Búsqueda Avanzada")
        self.geometry("400x200")
        self.resizable(False, False)
        self.grab_set()

        self.callback = callback
        
        # Mapping from display name to database column name
        self.search_fields = {
            "Código": "codigo",
            "País": "pais",
            "Ciudad": "ciudad",
            "Designación Tienda": "designacion_tienda",
            "Peticionario": "peticionario_name",
            "Versión Modelo Imagen": "image_model_version",
            "Versión Modelo Constructivo": "constructive_model_version",
            "Versión Mobiliario": "mobiliario_version_propuesta",
            "Versión Puertas": "puertas_version_propuesta"
        }

        main_frame = ctk.CTkFrame(self)
        main_frame.pack(expand=True, fill="both", padx=15, pady=15)
        main_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(main_frame, text="Buscar por campo:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.field_menu = ctk.CTkOptionMenu(main_frame, values=list(self.search_fields.keys()))
        self.field_menu.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(main_frame, text="Término:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.term_entry = ctk.CTkEntry(main_frame, placeholder_text="Escriba aquí...")
        self.term_entry.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
        self.term_entry.bind("<Return>", self.perform_search)

        search_button = ctk.CTkButton(main_frame, text="Buscar", command=self.perform_search, fg_color="green")
        search_button.grid(row=2, column=0, columnspan=2, pady=15, padx=10)
        
        self.term_entry.focus()

    def perform_search(self, event=None):
        search_term = self.term_entry.get().strip()
        if not search_term:
            messagebox.showwarning("Entrada Vacía", "Por favor, introduzca un término de búsqueda.", parent=self)
            return
        
        display_field = self.field_menu.get()
        db_field = self.search_fields[display_field]
        
        self.callback(field=db_field, term=search_term)
        self.destroy()

class App(ctk.CTk):
    def __init__(self, user_id, user_role):
        super().__init__()
        
        self.title("BIM_Manager")
        self.geometry("1800x900") # Increased default size
        self.resizable(True, True) # Allow resizing

        # Configure window icon
        try:
            # CORRECCIÓN: Usar una ruta absoluta o relativa robusta
            icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
            if os.path.exists(icon_path):
                self.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"Advertencia: No se pudo cargar 'icon.ico'. Asegúrese de que el archivo existe y es válido: {e}")

        # Handle window closing event
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.current_selected_project_id = None
        self.current_user_id = user_id
        self.current_user_role = user_role

        # --- Inicialización de variables de control (BooleanVar) ---
        self.vars = {
            "mobiliario_incorporado": ctk.BooleanVar(),
            "puertas_incorporadas": ctk.BooleanVar(),
            "image_model_date_today": ctk.BooleanVar(),
            "constructive_model_date_today": ctk.BooleanVar()
        }

        # Initialize buttons/frames here so they can be accessed globally in App
        self.delete_project_button = None
        self.search_projects_frame = None 
        self.project_list_frame = None 
        self.project_details_scroll_frame = None 
        self.export_import_frame = None 
        self.summary_frame = None 
        self.user_management_frame = None
        self.dashboard_frame = None

        # Define sortable columns and their display text components
        # (base_display_text, internal_column_key, data_type)
        # The indicator is added dynamically in load_summary_data
        self.sortable_column_definitions = [
            ("Código País Ciudad", "codigo_pais_ciudad", "string"),
            ("Inicio", "fecha_inicio_obra", "date"),
            ("Apertura", "fecha_apertura", "date"),
            ("MI", "image_model_version", "string"),
            ("MC", "constructive_model_version", "string"),
            ("V", "mobiliario_version_propuesta", "string"),
            ("Fecha", "mobiliario_actualizacion_date", "date"),
            ("V", "puertas_version_propuesta", "string"),
            ("Fecha", "puertas_actualizacion_date", "date"),
            ("Peticionario", "last_peticionario", "string"),
            ("Revisión", "revision_version", "numeric")
        ]

        # Initialize sort state
        self.current_sort_column_key = self.sortable_column_definitions[0][1] # Default to 'codigo_pais_ciudad'
        self.current_sort_order = "asc"

        # Map for internal sorting logic (db_index, type)
        # These indices correspond to the order of columns in the SQL SELECT statement in _get_summary_data
        # NOTE: The db_index for 'codigo_pais_ciudad' now refers to the index of 'p.id' as well,
        # so it's (1, 2, 3) for (codigo, pais, ciudad) and 0 for id.
        self.sortable_columns_map = {
            "codigo_pais_ciudad": {"db_index": (1, 2, 3), "type": "string"}, # (codigo, pais, ciudad)
            "fecha_inicio_obra": {"db_index": 4, "type": "date"},
            "fecha_apertura": {"db_index": 5, "type": "date"},
            "image_model_version": {"db_index": 6, "type": "string"},
            "constructive_model_version": {"db_index": 7, "type": "string"},
            "mobiliario_version_propuesta": {"db_index": 8, "type": "string"},
            "mobiliario_actualizacion_date": {"db_index": 9, "type": "date"},
            "puertas_version_propuesta": {"db_index": 10, "type": "string"},
            "puertas_actualizacion_date": {"db_index": 11, "type": "date"},
            "last_peticionario": {"db_index": 12, "type": "string"},
            "revision_version": {"db_index": 13, "type": "numeric"}
        }

        # --- Sidebar Frame ---
        self.create_sidebar()

        # --- Main Content Frame ---
        self.main_content_frame = ctk.CTkFrame(self, corner_radius=10)
        self.main_content_frame.grid(row=0, column=1, padx=(10, 20), pady=(10, 20), sticky="nsew")
        self.main_content_frame.grid_rowconfigure(0, weight=1)
        self.main_content_frame.grid_columnconfigure(0, weight=1)

        # --- View Frames (managed by dictionary) ---
        self.frames = {}
        # Select initial frame
        self.select_frame_by_name("summary")

    def create_sidebar(self):
        """
        Crea el frame lateral con los botones de navegación.
        """
        self.sidebar_frame = ctk.CTkFrame(self, width=180, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(9, weight=1) # Espacio flexible

        # Título de la App
        logo_label = ctk.CTkLabel(self.sidebar_frame, text="BIM Manager", font=ctk.CTkFont(size=24, weight="bold"))
        logo_label.grid(row=0, column=0, padx=20, pady=20)

        # Use a dictionary to store buttons for easier management
        self.buttons = {
            "dashboard": ctk.CTkButton(self.sidebar_frame, text="DASHBOARD",
                                            command=lambda: self.select_frame_by_name("dashboard"),
                                            fg_color="purple", height=60, font=ctk.CTkFont(weight="bold")),
            "summary": ctk.CTkButton(self.sidebar_frame, text="RESUMEN",
                                            command=lambda: self.select_frame_by_name("summary"),
                                            fg_color="green", height=60, font=ctk.CTkFont(weight="bold")),
            "registered_projects": ctk.CTkButton(self.sidebar_frame, text="PROYECTOS",
                                              command=lambda: self.select_frame_by_name("registered_projects"), font=ctk.CTkFont(weight="bold")),
            "new_project": ctk.CTkButton(self.sidebar_frame, text="NUEVO",
                                                  command=lambda: self.select_frame_by_name("new_project"), fg_color="red", font=ctk.CTkFont(weight="bold")),
            "search_projects": ctk.CTkButton(self.sidebar_frame, text="BUSCAR",
                                            command=lambda: self.select_frame_by_name("search_projects"), font=ctk.CTkFont(weight="bold")),
            "export_import": ctk.CTkButton(self.sidebar_frame, text="EXPORT / IMPORT",
                                                  command=lambda: self.select_frame_by_name("export_import"), font=ctk.CTkFont(weight="bold")),
            "settings": ctk.CTkButton(self.sidebar_frame, text="AJUSTES",
                                             command=lambda: self.select_frame_by_name("settings"), font=ctk.CTkFont(weight="bold")),
            "help": ctk.CTkButton(self.sidebar_frame, text="AYUDA",
                                         command=lambda: self.select_frame_by_name("help"), font=ctk.CTkFont(weight="bold")),
            "about": ctk.CTkButton(self.sidebar_frame, text="ACERCA DE",
                                          command=lambda: self.select_frame_by_name("about"), font=ctk.CTkFont(weight="bold")),
            "user_management": ctk.CTkButton(self.sidebar_frame, text="USUARIOS",
                                            command=lambda: self.select_frame_by_name("user_management"), font=ctk.CTkFont(weight="bold")),
            "close_app": ctk.CTkButton(self.sidebar_frame, text="CERRAR",
                                            command=self.on_closing, fg_color="#D32F2F", hover_color="#B71C1C", font=ctk.CTkFont(weight="bold"))
        }

        # Grid placement for buttons
        self.buttons["dashboard"].grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        self.buttons["summary"].grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        ctk.CTkFrame(self.sidebar_frame, height=2, fg_color="gray").grid(row=3, column=0, padx=20, pady=10, sticky="ew")
        self.buttons["registered_projects"].grid(row=4, column=0, padx=20, pady=10, sticky="ew")
        self.buttons["new_project"].grid(row=5, column=0, padx=20, pady=10, sticky="ew")
        self.buttons["search_projects"].grid(row=6, column=0, padx=20, pady=10, sticky="ew")
        self.buttons["export_import"].grid(row=7, column=0, padx=20, pady=10, sticky="ew")
        ctk.CTkFrame(self.sidebar_frame, height=2, fg_color="gray").grid(row=8, column=0, padx=20, pady=10, sticky="ew")
        
        # Bottom buttons
        self.buttons["settings"].grid(row=10, column=0, padx=20, pady=10, sticky="s")
        self.buttons["help"].grid(row=11, column=0, padx=20, pady=10, sticky="s")
        self.buttons["about"].grid(row=12, column=0, padx=20, pady=10, sticky="s")
        
        # Mostrar el botón de gestión de usuarios solo si el rol es 'Admin'
        if self.current_user_role == 'Admin':
            self.buttons["user_management"].grid(row=13, column=0, padx=20, pady=10, sticky="s")
        
        self.buttons["close_app"].grid(row=14, column=0, padx=20, pady=20, sticky="s")


    def on_closing(self):
        """Handles the window closing event."""
        logging.info("Aplicación cerrada por el usuario.")
        if messagebox.askokcancel("Salir", "¿Seguro que quieres salir?"):
            self.destroy()
            sys.exit(0) # Exit cleanly

    def select_frame_by_name(self, name):
        """
        Shows the frame corresponding to the name and hides the others.
        Applies role-based access control.
        """
        logging.info(f"Navegando a la pestaña: {name}")

        # Role-based access control
        if name in ["new_project", "export_import"] and self.current_user_role == 'Visitor':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para acceder a esta funcionalidad.")
            logging.warning(f"Intento de acceso denegado a '{name}' para el rol '{self.current_user_role}'.")
            return
        if name == "user_management" and self.current_user_role != 'Admin':
            messagebox.showwarning("Acceso Denegado", "Solo los administradores pueden acceder a la gestión de usuarios.")
            logging.warning(f"Intento de acceso denegado a 'user_management' para el rol '{self.current_user_role}'.")
            return
        
        # Get default button color from theme
        default_button_color = ctk.ThemeManager.theme["CTkButton"]["fg_color"]
        unselected_button_color = ("gray75", "gray25")

        # Reset all buttons to unselected color first
        for btn_name, button in self.buttons.items():
            if btn_name == "summary":
                button.configure(fg_color="green")
            elif btn_name == "dashboard":
                button.configure(fg_color="purple")
            elif btn_name == "new_project":
                button.configure(fg_color="red")
            elif btn_name == "close_app":
                button.configure(fg_color="#D32F2F")
            else:
                button.configure(fg_color=unselected_button_color)


        # Set the selected button to its specific color
        if name in self.buttons:
            if name not in ["summary", "dashboard", "new_project", "close_app"]:
                self.buttons[name].configure(fg_color=default_button_color)

        # Hide all frames
        for frame_widget in self.frames.values():
            frame_widget.grid_forget()

        # Create and show the selected frame
        if name not in self.frames:
            creator_method_name = f"create_{name}_frame"
            if hasattr(self, creator_method_name):
                creator_method = getattr(self, creator_method_name)
                self.frames[name] = creator_method(self.main_content_frame)
            else:
                logging.error(f"Método de creación de frame no encontrado: {creator_method_name}")
                messagebox.showerror("Error de Desarrollo", f"El método para crear el frame '{name}' no existe: {creator_method_name}")
                return 
        
        self.frames[name].grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # Perform specific actions for certain frames
        if name == "new_project":
            self._toggle_mobiliario_fields()
            self._toggle_puertas_fields()
        elif name == "registered_projects":
            self.load_project_list()
            if self.project_details_scroll_frame: 
                for widget in self.project_details_scroll_frame.winfo_children():
                    widget.destroy()
                ctk.CTkLabel(self.project_details_scroll_frame, 
                             text="Seleccione un proyecto de la lista para ver sus detalles.",
                             font=ctk.CTkFont(size=16, slant="italic", weight="bold")).pack(pady=50, padx=20, fill="both", expand=True)
            self.current_selected_project_id = None
            if self.delete_project_button:
                self.delete_project_button.configure(state="disabled")
            if self.current_user_role == 'Visitor':
                if self.delete_project_button:
                    self.delete_project_button.configure(state="disabled")
        elif name == "search_projects": 
            if hasattr(self, 'search_results_frame') and self.search_results_frame:
                for widget in self.search_results_frame.winfo_children():
                    widget.destroy()
                ctk.CTkLabel(self.search_results_frame, text="Introduzca un término de búsqueda global y pulse 'BUSCAR'.",
                             font=ctk.CTkFont(size=14, slant="italic", weight="bold")).pack(pady=20, padx=10)
        elif name == "summary": 
            self.load_summary_data() 
        elif name == "user_management":
            self.load_user_list()
        elif name == "dashboard":
            self._update_dashboard_chart()


    # --- Functions and Frames for NEW PROJECT ---
    def _set_date_to_today_for_new_project(self, entry_widget, var_checkbox):
        """Establece la fecha del campo de entrada a la fecha actual si el checkbox está marcado (para Nuevo Proyecto)."""
        if var_checkbox.get():
            entry_widget.delete(0, ctk.END)
            entry_widget.insert(0, datetime.now().strftime('%Y-%m-%d'))
        else:
            entry_widget.delete(0, ctk.END) # Limpiar si desmarca

    def _toggle_mobiliario_fields(self):
        """Habilita/deshabilita los campos de mobiliario según el checkbox (adaptado para la clase App)."""
        if hasattr(self, 'entry_mobiliario_fecha_propuesta'): 
            if self.vars["mobiliario_incorporado"].get(): # Use self.vars
                self.entry_mobiliario_fecha_propuesta.configure(state="normal")
                self.entry_mobiliario_version_propuesta.configure(state="normal")
                self.entry_mobiliario_incorporation_date.configure(state="normal")
            else:
                self.entry_mobiliario_fecha_propuesta.configure(state="disabled")
                self.entry_mobiliario_version_propuesta.configure(state="disabled")
                self.entry_mobiliario_incorporation_date.configure(state="disabled")
                self.entry_mobiliario_fecha_propuesta.delete(0, ctk.END)
                self.entry_mobiliario_version_propuesta.delete(0, ctk.END)
                self.entry_mobiliario_incorporation_date.delete(0, ctk.END)

    def _toggle_puertas_fields(self):
        """Habilita/deshabilita los campos de puertas según el checkbox (adaptado para la clase App)."""
        if hasattr(self, 'entry_puertas_fecha_propuesta'): 
            if self.vars["puertas_incorporadas"].get(): # Use self.vars
                self.entry_puertas_fecha_propuesta.configure(state="normal")
                self.entry_puertas_version_propuesta.configure(state="normal")
                self.entry_puertas_incorporation_date.configure(state="normal")
            else:
                self.entry_puertas_fecha_propuesta.configure(state="disabled")
                self.entry_puertas_version_propuesta.configure(state="disabled")
                self.entry_puertas_incorporation_date.configure(state="disabled")
                self.entry_puertas_fecha_propuesta.delete(0, ctk.END)
                self.entry_puertas_version_propuesta.delete(0, ctk.END)
                self.entry_puertas_incorporation_date.delete(0, ctk.END)

    def _save_project(self):
        """
        Guarda los datos del proyecto en la base de datos (adaptado para la clase App).
        """
        if self.current_user_role == 'Visitor':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para crear proyectos.")
            logging.warning(f"Intento de creación de proyecto denegado para el rol '{self.current_user_role}'.")
            return

        codigo = self.entry_codigo.get().strip()
        cadena = self.entry_cadena.get().strip()
        pais = self.entry_pais.get().strip()
        ciudad = self.entry_ciudad.get().strip()
        ubicacion = self.entry_ubicacion.get().strip()
        designacion_tienda = self.entry_designacion_tienda.get().strip()

        mobiliario_incorporado = 1 if self.vars["mobiliario_incorporado"].get() else 0 # Use self.vars
        mobiliario_fecha_propuesta = self.entry_mobiliario_fecha_propuesta.get().strip()
        mobiliario_version_propuesta = self.entry_mobiliario_version_propuesta.get().strip()
        mobiliario_actualizacion_date = "" 
        mobiliario_incorporation_date = self.entry_mobiliario_incorporation_date.get().strip()

        puertas_incorporadas = 1 if self.vars["puertas_incorporadas"].get() else 0 # Use self.vars
        puertas_fecha_propuesta = self.entry_puertas_fecha_propuesta.get().strip()
        puertas_version_propuesta = self.entry_puertas_version_propuesta.get().strip()
        puertas_actualizacion_date = "" 
        puertas_incorporation_date = self.entry_puertas_incorporation_date.get().strip()

        fecha_inicio_obra = self.entry_fecha_inicio_obra.get().strip()
        fecha_apertura = self.entry_fecha_apertura.get().strip()

        image_model_version = self.entry_image_model_version.get().strip()
        image_model_date = self.entry_image_model_date.get().strip() # Obtener de la entrada
        constructive_model_version = self.entry_constructive_model_version.get().strip()
        constructive_model_date = self.entry_constructive_model_date.get().strip() # Obtener de la entrada

        creation_date = datetime.now().strftime('%Y-%m-%d') # Fecha de creación automática
        revision_version = 0 # Inicializar la versión de revisión a 0 para nuevos proyectos


        if not all([codigo, cadena, pais, ciudad, ubicacion, designacion_tienda]):
            messagebox.showerror("Error de Entrada", "Por favor, complete todos los campos de la nomenclatura del proyecto.")
            return

        nombre_completo = f"{codigo} {cadena} {pais} {ciudad} {ubicacion} ({designacion_tienda})"

        dates_to_validate = [
            image_model_date, constructive_model_date,
            mobiliario_fecha_propuesta, mobiliario_incorporation_date,
            puertas_fecha_propuesta, puertas_incorporation_date,
            fecha_inicio_obra, fecha_apertura
        ]
        for date_str in dates_to_validate:
            if date_str and not is_valid_date(date_str):
                messagebox.showerror("Error de Fecha", f"La fecha '{date_str}' no tiene un formato válido (YYYY-MM-DD).")
                return

        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO projects (
                    codigo, cadena, pais, ciudad, ubicacion, designacion_tienda, nombre_completo,
                    image_model_version, image_model_date, constructive_model_version, constructive_model_date,
                    mobiliario_incorporado, mobiliario_fecha_propuesta, mobiliario_version_propuesta,
                    mobiliario_actualizacion_date, mobiliario_incorporation_date,
                    puertas_incorporadas, puertas_fecha_propuesta, puertas_version_propuesta,
                    puertas_actualizacion_date, puertas_incorporation_date,
                    fecha_inicio_obra, fecha_apertura,
                    creation_date, revision_version
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                codigo, cadena, pais, ciudad, ubicacion, designacion_tienda, nombre_completo,
                image_model_version, image_model_date, constructive_model_version, constructive_model_date,
                mobiliario_incorporado, mobiliario_fecha_propuesta, mobiliario_version_propuesta,
                mobiliario_actualizacion_date, mobiliario_incorporation_date,
                puertas_incorporadas, puertas_fecha_propuesta, puertas_version_propuesta,
                puertas_actualizacion_date, puertas_incorporation_date,
                fecha_inicio_obra, fecha_apertura,
                creation_date, revision_version
            ))
            project_id = cursor.lastrowid # Get the ID of the newly inserted project

            # Log the creation as a history entry
            # For new project creation, Peticionario can be the current user's username
            conn_user = sqlite3.connect(DB_NAME)
            cursor_user = conn_user.cursor()
            cursor_user.execute("SELECT username FROM users WHERE id = ?", (self.current_user_id,))
            current_username_tuple = cursor_user.fetchone()
            current_username = current_username_tuple[0] if current_username_tuple else "Sistema"
            conn_user.close()

            cursor.execute('''
                INSERT INTO project_updates_history (project_id, user_id, peticionario_name, revision_version_at_update, update_date, changed_field, old_value, new_value)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (project_id, self.current_user_id, current_username, 0, creation_date, "Proyecto Creado", "N/A", nombre_completo))

            conn.commit()
            logging.info(f"Proyecto '{nombre_completo}' guardado correctamente por usuario ID {self.current_user_id}. ID: {project_id}")
            messagebox.showinfo("Éxito", f"Proyecto '{nombre_completo}' guardado correctamente.")
            self._clear_entries()
        except sqlite3.Error as e:
            logging.error(f"Error de base de datos al guardar proyecto: {e}")
            messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al guardar el proyecto: {e}")
        finally:
            if conn:
                conn.close()

    def _clear_entries(self):
        """
        Limpia todos los campos de entrada del formulario (adaptado para la clase App).
        """
        self.entry_codigo.delete(0, ctk.END)
        self.entry_cadena.delete(0, ctk.END)
        self.entry_pais.delete(0, ctk.END)
        self.entry_ciudad.delete(0, ctk.END)
        self.entry_ubicacion.delete(0, ctk.END)
        self.entry_designacion_tienda.delete(0, ctk.END)
        
        self.entry_image_model_version.delete(0, ctk.END)
        self.entry_image_model_date.delete(0, ctk.END)
        self.vars["image_model_date_today"].set(False) # Desmarcar checkbox
        
        self.entry_constructive_model_version.delete(0, ctk.END)
        self.entry_constructive_model_date.delete(0, ctk.END)
        self.vars["constructive_model_date_today"].set(False) # Desmarcar checkbox

        self.vars["mobiliario_incorporado"].set(False)
        self.entry_mobiliario_fecha_propuesta.delete(0, ctk.END)
        self.entry_mobiliario_version_propuesta.delete(0, ctk.END)
        self.entry_mobiliario_incorporation_date.delete(0, ctk.END)
        
        self.vars["puertas_incorporadas"].set(False)
        self.entry_puertas_fecha_propuesta.delete(0, ctk.END)
        self.entry_puertas_version_propuesta.delete(0, ctk.END)
        self.entry_puertas_incorporation_date.delete(0, ctk.END)
        
        self.entry_fecha_inicio_obra.delete(0, ctk.END)
        self.entry_fecha_apertura.delete(0, ctk.END)
        

        self._toggle_mobiliario_fields()
        self._toggle_puertas_fields()

    def create_new_project_frame(self, parent_frame):
        """
        Creates and returns the frame for entering new projects.
        """
        form_frame = ctk.CTkScrollableFrame(parent_frame, corner_radius=10) 
        form_frame.grid_columnconfigure(0, weight=1, uniform="nomen_col") 
        form_frame.grid_columnconfigure(1, weight=1, uniform="nomen_col") 
        form_frame.grid_columnconfigure(2, weight=1, uniform="nomen_col") 
        form_frame.grid_columnconfigure(3, weight=1, uniform="nomen_col") 
        form_frame.grid_columnconfigure(4, weight=1, uniform="nomen_col") 
        form_frame.grid_columnconfigure(5, weight=1, uniform="nomen_col") 

        ctk.CTkLabel(form_frame, text="Detalles del Nuevo Proyecto", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, columnspan=6, pady=(0, 15))

        # --- Inicializar CTkEntry aquí, con form_frame como padre ---
        self.entry_codigo = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_cadena = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_pais = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_ciudad = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_ubicacion = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_designacion_tienda = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        
        self.entry_image_model_version = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_image_model_date = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold")) # Nueva entrada para fecha
        
        self.entry_constructive_model_version = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_constructive_model_date = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold")) # Nueva entrada para fecha
        
        self.entry_mobiliario_fecha_propuesta = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold"))
        self.entry_mobiliario_version_propuesta = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_mobiliario_incorporation_date = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold"))

        self.entry_puertas_fecha_propuesta = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold"))
        self.entry_puertas_version_propuesta = ctk.CTkEntry(form_frame, font=ctk.CTkFont(weight="bold"))
        self.entry_puertas_incorporation_date = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold"))

        self.entry_fecha_inicio_obra = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold"))
        self.entry_fecha_apertura = ctk.CTkEntry(form_frame, placeholder_text="YYYY-MM-DD", font=ctk.CTkFont(weight="bold"))
        # --- Fin de inicialización de CTkEntry ---


        # --- Sección de Nomenclatura (Formato Tabla) ---
        nomenclatura_label = ctk.CTkLabel(form_frame, text="--- Nomenclatura del Proyecto ---", font=ctk.CTkFont(size=16, weight="bold", underline=True))
        nomenclatura_label.grid(row=1, column=0, columnspan=6, sticky="w", pady=(10, 5), padx=5)

        header_font = ctk.CTkFont(size=13, weight="bold", underline=True)
        header_bg_color = "#2D68C4" 
        
        # Encabezados de la tabla de nomenclatura
        ctk.CTkLabel(form_frame, text="Código", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=2, column=0, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="Cadena", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=2, column=1, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="País", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=2, column=2, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="Ciudad", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=2, column=3, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="Ubicación", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=2, column=4, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="Designación Tienda", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=2, column=5, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)

        # Campos de entrada de la nomenclatura
        self.entry_codigo.grid(row=3, column=0, sticky="ew", padx=2, pady=2)
        self.entry_cadena.grid(row=3, column=1, sticky="ew", padx=2, pady=2)
        self.entry_pais.grid(row=3, column=2, sticky="ew", padx=2, pady=2)
        self.entry_ciudad.grid(row=3, column=3, sticky="ew", padx=2, pady=2)
        self.entry_ubicacion.grid(row=3, column=4, sticky="ew", padx=2, pady=2)
        self.entry_designacion_tienda.grid(row=3, column=5, sticky="ew", padx=2, pady=2)

        current_row = 4 

        # --- Separador ---
        ctk.CTkLabel(form_frame, text="").grid(row=current_row, columnspan=6, pady=5)
        current_row += 1

        # --- Matriz de Modelos, Mobiliario, Puertas ---
        matrix_section_label = ctk.CTkLabel(form_frame, text="--- Modelos y Elementos ---", font=ctk.CTkFont(size=16, weight="bold", underline=True))
        matrix_section_label.grid(row=current_row, column=0, columnspan=6, sticky="w", pady=(10, 5), padx=5)
        current_row += 1

        # Encabezados de la Matriz - Resaltados con fondo azul
        ctk.CTkLabel(form_frame, text="TIPO DE MODELO / ELEMENTO", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=current_row, column=0, columnspan=2, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5) 
        ctk.CTkLabel(form_frame, text="VERSIÓN", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=current_row, column=2, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="FECHA PROPUESTA", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=current_row, column=3, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
        ctk.CTkLabel(form_frame, text="1º INCORPORACION", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=current_row, column=4, columnspan=2, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5) 
        current_row += 1

        # Modelo Imagen
        ctk.CTkLabel(form_frame, text="Modelo Imagen (MI)", font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=0, columnspan=2, sticky="w", padx=5, pady=2)
        self.entry_image_model_version.grid(row=current_row, column=2, sticky="ew", padx=5, pady=2)
        self.entry_image_model_date.grid(row=current_row, column=3, sticky="ew", padx=5, pady=2) # Campo de entrada para la fecha
        ctk.CTkCheckBox(form_frame, text="Fecha Hoy", variable=self.vars["image_model_date_today"], # Use self.vars
                        command=lambda e=self.entry_image_model_date, v=self.vars["image_model_date_today"]: self._set_date_to_today_for_new_project(e, v), font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=4, sticky="w", padx=5)
        ctk.CTkLabel(form_frame, text="N/A", font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=5, sticky="w", padx=5, pady=2) # Columna 5 para N/A
        current_row += 1

        # Modelo Constructivo
        ctk.CTkLabel(form_frame, text="Modelo Constructivo (MC)", font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=0, columnspan=2, sticky="w", padx=5, pady=2)
        self.entry_constructive_model_version.grid(row=current_row, column=2, sticky="ew", padx=5, pady=2)
        self.entry_constructive_model_date.grid(row=current_row, column=3, sticky="ew", padx=5, pady=2) # Campo de entrada para la fecha
        ctk.CTkCheckBox(form_frame, text="Fecha Hoy", variable=self.vars["constructive_model_date_today"], # Use self.vars
                        command=lambda e=self.entry_constructive_model_date, v=self.vars["constructive_model_date_today"]: self._set_date_to_today_for_new_project(e, v), font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=4, sticky="w", padx=5)
        ctk.CTkLabel(form_frame, text="N/A", font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=5, sticky="w", padx=5, pady=2) # Columna 5 para N/A
        current_row += 1

        # Mobiliario (150 MOA)
        ctk.CTkCheckBox(form_frame, text="MOBILIARIO (150 MOA)", variable=self.vars["mobiliario_incorporado"], command=self._toggle_mobiliario_fields, font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=0, columnspan=2, sticky="w", padx=5, pady=5)
        self.entry_mobiliario_version_propuesta.grid(row=current_row, column=2, sticky="ew", padx=5, pady=2)
        self.entry_mobiliario_fecha_propuesta.grid(row=current_row, column=3, sticky="ew", padx=5, pady=2)
        self.entry_mobiliario_incorporation_date.grid(row=current_row, column=4, columnspan=2, sticky="ew", padx=5, pady=2)
        current_row += 1

        # Puertas (170 CMD)
        ctk.CTkCheckBox(form_frame, text="PUERTAS (170 CMD)", variable=self.vars["puertas_incorporadas"], command=self._toggle_puertas_fields, font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=0, columnspan=2, sticky="w", padx=5, pady=5)
        self.entry_puertas_version_propuesta.grid(row=current_row, column=2, sticky="ew", padx=5, pady=2)
        self.entry_puertas_fecha_propuesta.grid(row=current_row, column=3, sticky="ew", padx=5, pady=2)
        self.entry_puertas_incorporation_date.grid(row=current_row, column=4, columnspan=2, sticky="ew", padx=5, pady=2)
        current_row += 1

        # --- Separador ---
        ctk.CTkLabel(form_frame, text="").grid(row=current_row, columnspan=6, pady=5)
        current_row += 1

        # --- Sección de Fechas de Obra ---
        obra_label = ctk.CTkLabel(form_frame, text="--- Fechas de Obra ---", font=ctk.CTkFont(size=16, weight="bold", underline=True))
        obra_label.grid(row=current_row, column=0, columnspan=6, sticky="w", pady=(10, 5), padx=5)
        current_row += 1

        fields_obra = [
            ("Fecha Inicio Obra:", self.entry_fecha_inicio_obra),
            ("Fecha Apertura:", self.entry_fecha_apertura),
        ]
        for i, (text, entry_widget) in enumerate(fields_obra):
            ctk.CTkLabel(form_frame, text=text, font=ctk.CTkFont(weight="bold")).grid(row=current_row + i, column=0, columnspan=2, sticky="w", pady=2, padx=5)
            entry_widget.grid(row=current_row + i, column=2, columnspan=4, sticky="ew", pady=2, padx=5) 
        current_row += len(fields_obra)
        
        # Versión Revisión para Nuevo Proyecto (solo muestra el valor inicial)
        ctk.CTkLabel(form_frame, text="Versión Revisión: 0 (se actualiza al guardar)", font=ctk.CTkFont(weight="bold")).grid(row=current_row, column=0, columnspan=6, sticky="w", padx=5, pady=10)
        current_row += 1

        # Botón Guardar
        btn_save = ctk.CTkButton(form_frame, text="Guardar Proyecto", command=self._save_project, font=ctk.CTkFont(weight="bold"))
        btn_save.grid(row=current_row, column=0, columnspan=6, pady=20)

        return form_frame

    # --- Functions and Frames for REGISTERED PROJECTS ---
    def create_registered_projects_frame(self, parent_frame):
        """
        Creates and returns the frame for viewing registered projects.
        """
        registered_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        registered_frame.grid_columnconfigure(1, weight=1) 
        registered_frame.grid_rowconfigure(0, weight=1)

        # Frame for the project list (left)
        self.project_list_container_frame = ctk.CTkFrame(registered_frame, corner_radius=10)
        self.project_list_container_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        self.project_list_container_frame.grid_rowconfigure(1, weight=1) 
        self.project_list_container_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.project_list_container_frame, text="PROYECTOS", font=ctk.CTkFont(size=18, weight="bold")).grid(row=0, column=0, pady=(10, 5))

        self.project_list_frame = ctk.CTkScrollableFrame(self.project_list_container_frame, width=250, height=500, corner_radius=10)
        self.project_list_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.project_list_frame.grid_columnconfigure(0, weight=1)

        # Botón para eliminar proyecto
        self.delete_project_button = ctk.CTkButton(self.project_list_container_frame, text="Eliminar Proyecto",
                                                    command=self._delete_selected_project, fg_color="red", state="disabled", font=ctk.CTkFont(weight="bold"))
        self.delete_project_button.grid(row=2, column=0, padx=20, pady=10)


        # Frame for the selected project details (right)
        self.project_details_scroll_frame = ctk.CTkScrollableFrame(registered_frame, corner_radius=10)
        self.project_details_scroll_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        self.project_details_scroll_frame.grid_columnconfigure(0, weight=1) 

        # Initial label for when no project is selected
        self.initial_details_label = ctk.CTkLabel(self.project_details_scroll_frame, 
                                                  text="Seleccione un proyecto de la lista para ver sus detalles.",
                                                  font=ctk.CTkFont(size=16, slant="italic", weight="bold"))
        self.initial_details_label.pack(pady=50, padx=20, fill="both", expand=True)

        return registered_frame

    def load_project_list(self):
        """
        Loads the project list into the left frame.
        """
        # Clear existing list
        for widget in self.project_list_frame.winfo_children():
            widget.destroy()

        projects = get_all_projects()
        if not projects:
            ctk.CTkLabel(self.project_list_frame, text="No hay proyectos registrados.", font=ctk.CTkFont(weight="bold")).pack(pady=10) # Negrita
            if self.delete_project_button:
                self.delete_project_button.configure(state="disabled")
            self.current_selected_project_id = None
            return

        for project_id, codigo, ciudad, nombre_completo in projects:
            btn = ctk.CTkButton(self.project_list_frame, text=f"{codigo} {ciudad}", 
                                command=lambda p_id=project_id: self.display_project_details(p_id), font=ctk.CTkFont(weight="bold")) # Negrita
            btn.pack(fill="x", pady=5, padx=5)
        
        # Disable delete button and details editing for Visitors
        if self.current_user_role == 'Visitor':
            if self.delete_project_button:
                self.delete_project_button.configure(state="disabled")
        else:
            if self.delete_project_button:
                self.delete_project_button.configure(state="disabled") # Initially disabled until a project is selected
        self.current_selected_project_id = None


    def display_project_details(self, project_id):
        """
        Displays the details of a specific project in the details area with an elegant table design.
        """
        project = get_project_details(project_id)
        self.current_selected_project_id = project_id 

        # Enable delete button if user is Admin or Editor
        if self.current_user_role in ['Admin', 'Editor']:
            if self.delete_project_button:
                self.delete_project_button.configure(state="normal")
        else:
            if self.delete_project_button:
                self.delete_project_button.configure(state="disabled")

        for widget in self.project_details_scroll_frame.winfo_children():
            widget.destroy()

        if project:
            header_frame = ctk.CTkFrame(self.project_details_scroll_frame, fg_color="transparent")
            header_frame.pack(fill="x", pady=(10, 20), padx=15)
            header_frame.grid_columnconfigure(0, weight=1) 

            full_name = project.get('nombre_completo', 'N/A')
            ctk.CTkLabel(header_frame, text=full_name, 
                         font=ctk.CTkFont(size=22, weight="bold")).grid(row=0, column=0, sticky="w")
            
            btn_update_data = ctk.CTkButton(header_frame, text="Actualizar Datos del Proyecto", 
                                            command=lambda: self._open_update_dialog(project), font=ctk.CTkFont(weight="bold")) # Negrita
            
            # Disable update button for Visitors
            if self.current_user_role == 'Visitor':
                btn_update_data.configure(state="disabled")
            
            btn_update_data.grid(row=0, column=1, sticky="e", padx=(20,0))

            # --- Main frame for the details table (reverted to v17 style but with enhanced labels) ---
            details_table_frame = ctk.CTkFrame(self.project_details_scroll_frame, corner_radius=10, fg_color=("gray80", "gray20"))
            details_table_frame.pack(fill="both", padx=15, pady=10, expand=True)
            
            # --- Información General del Proyecto (Nueva Sección) ---
            general_info_frame = ctk.CTkFrame(details_table_frame, fg_color="transparent")
            general_info_frame.pack(fill="x", padx=10, pady=(10, 5))
            general_info_frame.grid_columnconfigure(0, weight=1)
            general_info_frame.grid_columnconfigure(1, weight=2)

            ctk.CTkLabel(general_info_frame, text="--- Información General ---", font=ctk.CTkFont(size=16, weight="bold", underline=True)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 5))
            
            # Mostrar la Fecha de Alta del Proyecto
            ctk.CTkLabel(general_info_frame, text="Fecha de Alta del Proyecto:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, sticky="w", padx=5, pady=2)
            ctk.CTkLabel(general_info_frame, text=project.get('creation_date', 'N/A'), font=ctk.CTkFont(weight="bold")).grid(row=1, column=1, sticky="w", padx=5, pady=2) # Negrita

            # Mostrar la Versión Revisión
            ctk.CTkLabel(general_info_frame, text="Versión Revisión:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, sticky="w", padx=5, pady=2)
            ctk.CTkLabel(general_info_frame, text=str(project.get('revision_version', 0)), font=ctk.CTkFont(weight="bold")).grid(row=2, column=1, sticky="w", padx=5, pady=2) # Negrita


            # --- Modelos y Elementos (Matrix-like presentation) ---
            matrix_frame = ctk.CTkFrame(details_table_frame, fg_color="transparent")
            matrix_frame.pack(fill="both", padx=10, pady=(10, 5), expand=True)
            
            matrix_frame.grid_columnconfigure(0, weight=6, uniform="matrix_col") 
            matrix_frame.grid_columnconfigure(1, weight=3, uniform="matrix_col") 
            matrix_frame.grid_columnconfigure(2, weight=4, uniform="matrix_col") 
            matrix_frame.grid_columnconfigure(3, weight=4, uniform="matrix_col") 

            header_font = ctk.CTkFont(size=13, weight="bold", underline=True)
            header_bg_color = "#2D68C4" 

            ctk.CTkLabel(matrix_frame, text="TIPO DE MODELO / ELEMENTO", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=0, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
            ctk.CTkLabel(matrix_frame, text="VERSIÓN", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=1, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
            ctk.CTkLabel(matrix_frame, text="ACTUALIZACIÓN", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=2, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
            ctk.CTkLabel(matrix_frame, text="1º INCORPORACION", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=3, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)

            current_row_matrix = 1

            # Data for Models (sin Revit)
            model_data = [
                ("Modelo Imagen (MI)", project.get('image_model_version', 'N/A'), project.get('image_model_date', 'N/A'), "N/A"),
                ("Modelo Constructivo (MC)", project.get('constructive_model_version', 'N/A'), project.get('constructive_model_date', 'N/A'), "N/A"),
                ("MOBILIARIO (150 MOA)", project.get('mobiliario_version_propuesta', 'N/A'), project.get('mobiliario_actualizacion_date', 'N/A'), project.get('mobiliario_incorporation_date', 'N/A')),
                ("PUERTAS (170 CMD)", project.get('puertas_version_propuesta', 'N/A'), project.get('puertas_actualizacion_date', 'N/A'), project.get('puertas_incorporation_date', 'N/A')),
            ]
            for item_type, version, update_date, first_inc_date in model_data:
                ctk.CTkLabel(matrix_frame, text=item_type, font=ctk.CTkFont(weight="bold")).grid(row=current_row_matrix, column=0, sticky="w", padx=5, pady=2)
                ctk.CTkLabel(matrix_frame, text=str(version), font=ctk.CTkFont(weight="bold")).grid(row=current_row_matrix, column=1, sticky="w", padx=5, pady=2) # Negrita
                ctk.CTkLabel(matrix_frame, text=str(update_date), font=ctk.CTkFont(weight="bold")).grid(row=current_row_matrix, column=2, sticky="w", padx=5, pady=2) # Negrita
                ctk.CTkLabel(matrix_frame, text=str(first_inc_date), font=ctk.CTkFont(weight="bold")).grid(row=current_row_matrix, column=3, sticky="w", padx=5, pady=2) # Negrita
                current_row_matrix += 1
            
            # --- Fechas de Obra Section ---
            obra_frame = ctk.CTkFrame(details_table_frame, fg_color="transparent")
            obra_frame.pack(fill="x", padx=10, pady=(10, 5))
            obra_frame.grid_columnconfigure(0, weight=1)
            obra_frame.grid_columnconfigure(1, weight=2)

            ctk.CTkLabel(obra_frame, text="\n--- Fechas de Obra ---", font=ctk.CTkFont(size=16, weight="bold", underline=True)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 5))
            data_obra = [
                ("Inicio Obra", project.get('fecha_inicio_obra', 'N/A')),
                ("Apertura", project.get('fecha_apertura', 'N/A')),
            ]
            for i, (label_text, value) in enumerate(data_obra):
                ctk.CTkLabel(obra_frame, text=f"{label_text}:", font=ctk.CTkFont(weight="bold")).grid(row=i+1, column=0, sticky="w", padx=5, pady=2)
                ctk.CTkLabel(obra_frame, text=str(value), font=ctk.CTkFont(weight="bold")).grid(row=i+1, column=1, sticky="w", padx=5, pady=2) # Negrita

            # --- Historial de Actualizaciones ---
            history_frame = ctk.CTkFrame(details_table_frame, fg_color="transparent")
            history_frame.pack(fill="x", padx=10, pady=(10, 5))
            history_frame.grid_columnconfigure(0, weight=1) # Usuario
            history_frame.grid_columnconfigure(1, weight=1) # Peticionario
            history_frame.grid_columnconfigure(2, weight=1) # Versión
            history_frame.grid_columnconfigure(3, weight=1) # Fecha
            history_frame.grid_columnconfigure(4, weight=1) # Campo
            history_frame.grid_columnconfigure(5, weight=1) # Old Value
            history_frame.grid_columnconfigure(6, weight=1) # New Value

            ctk.CTkLabel(history_frame, text="\n--- Historial de Actualizaciones ---", font=ctk.CTkFont(size=16, weight="bold", underline=True)).grid(row=0, column=0, columnspan=7, sticky="w", pady=(0, 5))
            
            history_data = get_project_update_history(project_id)

            if history_data:
                # Encabezados de la tabla de historial
                header_font = ctk.CTkFont(size=13, weight="bold", underline=True)
                header_bg_color = "#2D68C4"
                ctk.CTkLabel(history_frame, text="USUARIO APP", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=0, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
                ctk.CTkLabel(history_frame, text="PETICIONARIO", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=1, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
                ctk.CTkLabel(history_frame, text="VERSIÓN", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=2, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
                ctk.CTkLabel(history_frame, text="FECHA", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=3, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
                ctk.CTkLabel(history_frame, text="CAMPO", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=4, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
                ctk.CTkLabel(history_frame, text="VALOR ANTERIOR", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=5, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)
                ctk.CTkLabel(history_frame, text="NUEVO VALOR", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=1, column=6, sticky="nsew", padx=2, pady=2, ipadx=5, ipady=5)


                for i, (username, peticionario_name, version, date, changed_field, old_value, new_value) in enumerate(history_data):
                    ctk.CTkLabel(history_frame, text=username or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=0, sticky="w", padx=5, pady=2)
                    ctk.CTkLabel(history_frame, text=peticionario_name or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=1, sticky="w", padx=5, pady=2)
                    ctk.CTkLabel(history_frame, text=str(version), font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=2, sticky="w", padx=5, pady=2)
                    ctk.CTkLabel(history_frame, text=date, font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=3, sticky="w", padx=5, pady=2)
                    ctk.CTkLabel(history_frame, text=changed_field or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=4, sticky="w", padx=5, pady=2)
                    ctk.CTkLabel(history_frame, text=old_value or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=5, sticky="w", padx=5, pady=2)
                    ctk.CTkLabel(history_frame, text=new_value or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i+2, column=6, sticky="w", padx=5, pady=2)
            else:
                ctk.CTkLabel(history_frame, text="No hay historial de actualizaciones para este proyecto.", font=ctk.CTkFont(size=12, slant="italic", weight="bold")).grid(row=1, column=0, columnspan=7, pady=10) # Negrita


        else:
            ctk.CTkLabel(self.project_details_scroll_frame, text="No se encontraron detalles para este proyecto.",
                         font=ctk.CTkFont(size=16, slant="italic", weight="bold")).pack(pady=50, padx=20, fill="both", expand=True) # Negrita

    def _open_update_dialog(self, project_data):
        """
        Opens the dialog to update project data.
        """
        if self.current_user_role == 'Visitor':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para actualizar proyectos.")
            logging.warning(f"Intento de actualización de proyecto denegado para el rol '{self.current_user_role}'.")
            return
        UpdateProjectDialog(self, project_data, self.current_user_id, self.display_project_details)

    def _delete_selected_project(self):
        """
        Elimina el proyecto seleccionado de la base de datos.
        """
        if self.current_user_role not in ['Admin', 'Editor']:
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para eliminar proyectos.")
            logging.warning(f"Intento de eliminación de proyecto denegado para el rol '{self.current_user_role}'.")
            return

        if self.current_selected_project_id is None:
            messagebox.showinfo("Error", "Por favor, seleccione un proyecto para eliminar.")
            return

        project_details = get_project_details(self.current_selected_project_id)
        project_name = project_details.get('nombre_completo', 'este proyecto')

        if messagebox.askyesno("Confirmar Eliminación", f"¿Está seguro de que desea eliminar el proyecto '{project_name}'? Esta acción es irreversible."):
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            try:
                # La eliminación en cascada en la FK de project_updates_history se encargará de los registros relacionados
                cursor.execute("DELETE FROM projects WHERE id = ?", (self.current_selected_project_id,))
                conn.commit()
                logging.info(f"Proyecto '{project_name}' (ID: {self.current_selected_project_id}) eliminado correctamente por usuario ID {self.current_user_id}.")
                messagebox.showinfo("Éxito", f"Proyecto '{project_name}' eliminado correctamente.")
                self.load_project_list() 
                if self.project_details_scroll_frame: 
                    for widget in self.project_details_scroll_frame.winfo_children():
                        widget.destroy()
                    ctk.CTkLabel(self.project_details_scroll_frame,
                                                              text="Seleccione un proyecto de la lista para ver sus detalles.",
                                                              font=ctk.CTkFont(size=16, slant="italic", weight="bold")).pack(pady=50, padx=20, fill="both", expand=True) # Negrita
                self.current_selected_project_id = None
                if self.delete_project_button:
                    self.delete_project_button.configure(state="disabled") 
            except sqlite3.Error as e:
                logging.error(f"Error de base de datos al eliminar proyecto {self.current_selected_project_id}: {e}")
                messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al eliminar el proyecto: {e}")
            finally:
                if conn:
                    conn.close()

    def create_search_projects_frame(self, parent_frame):
        """
        Crea y retorna el frame para la funcionalidad de búsqueda de proyectos.
        """
        search_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        search_frame.grid_columnconfigure(0, weight=1)
        search_frame.grid_rowconfigure(2, weight=1)

        # Frame para los controles de búsqueda
        controls_frame = ctk.CTkFrame(search_frame)
        controls_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        controls_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(search_frame, text="Búsqueda de Proyectos", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, pady=(10, 5))

        # --- Campo de entrada para la búsqueda global ---
        self.global_search_entry = ctk.CTkEntry(controls_frame, placeholder_text="Buscar en todos los campos...")
        self.global_search_entry.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.global_search_entry.bind("<Return>", lambda event: self._perform_global_search())

        # Botón para ejecutar la búsqueda global
        self.search_execute_button = ctk.CTkButton(controls_frame, text="BUSCAR", command=self._perform_global_search, fg_color="green")
        self.search_execute_button.grid(row=0, column=1, padx=5)

        # Botón para búsqueda avanzada
        self.advanced_search_button = ctk.CTkButton(controls_frame, text="BÚSQUEDA AVANZADA", command=self._open_advanced_search, fg_color="green")
        self.advanced_search_button.grid(row=0, column=2, padx=5)

        # --- Marco para mostrar los resultados de la búsqueda ---
        self.search_results_frame = ctk.CTkScrollableFrame(search_frame, corner_radius=10, fg_color=("gray80", "gray20"))
        self.search_results_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        
        ctk.CTkLabel(self.search_results_frame, text="Introduzca un término de búsqueda y pulse 'BUSCAR' o use la Búsqueda Avanzada.",
                     font=ctk.CTkFont(size=14, slant="italic")).pack(pady=20, padx=10)

        return search_frame
    
    def _open_advanced_search(self):
        """Opens the advanced search dialog."""
        AdvancedSearchDialog(self, callback=self._perform_advanced_search)

    def _perform_advanced_search(self, field, term):
        """Performs a search on a specific field and displays results."""
        logging.info(f"Búsqueda avanzada realizada en el campo '{field}' para el término: '{term}'.")
        results = search_projects_in_db(advanced_field=field, advanced_term=term)
        self._display_search_results(results)

    def _perform_global_search(self):
        """
        Ejecuta la búsqueda global de proyectos basada en el término introducido
        y muestra los resultados en el search_results_frame.
        """
        global_search_term = self.global_search_entry.get().strip()
        logging.info(f"Búsqueda global realizada para el término: '{global_search_term}'.")
        results = search_projects_in_db(global_search_term=global_search_term)
        self._display_search_results(results)

    def _display_search_results(self, results):
        """Clears and populates the search results frame."""
        for widget in self.search_results_frame.winfo_children():
            widget.destroy()

        # Configurar las columnas en el frame de resultados para que se alineen correctamente
        self.search_results_frame.grid_columnconfigure(0, weight=3) # Proyecto
        self.search_results_frame.grid_columnconfigure(1, weight=1) # MI
        self.search_results_frame.grid_columnconfigure(2, weight=1) # Mobiliario
        self.search_results_frame.grid_columnconfigure(3, weight=1) # Puertas
        self.search_results_frame.grid_columnconfigure(4, weight=1) # Inicio Obra
        self.search_results_frame.grid_columnconfigure(5, weight=1) # Apertura

        if not results:
            ctk.CTkLabel(self.search_results_frame, text="No se encontraron proyectos con los criterios especificados.",
                         font=ctk.CTkFont(size=14, slant="italic", weight="bold")).grid(row=0, column=0, columnspan=6, pady=20, padx=10)
            return

        header_font = ctk.CTkFont(size=13, weight="bold", underline=True)
        header_bg_color = "#2D68C4"
        
        # New headers
        ctk.CTkLabel(self.search_results_frame, text="PROYECTOS", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.search_results_frame, text="MI", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=1, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.search_results_frame, text="Mobiliario", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=2, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.search_results_frame, text="Puertas", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=3, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.search_results_frame, text="Inicio Obra", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=4, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.search_results_frame, text="Apertura", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=5, sticky="nsew", padx=1, pady=1)

        for i, (project_id, full_name, mi_version, mob_version, puertas_version, fecha_inicio, fecha_apertura) in enumerate(results):
            # Button for project name
            btn = ctk.CTkButton(self.search_results_frame, text=full_name, anchor="w", fg_color="transparent", text_color_disabled=("gray", "gray"),
                          command=lambda p_id=project_id: (self.select_frame_by_name("registered_projects"), self.display_project_details(p_id)), font=ctk.CTkFont(weight="bold"))
            btn.grid(row=i + 1, column=0, sticky="ew", padx=5)
            
            # Other details
            ctk.CTkLabel(self.search_results_frame, text=mi_version or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i + 1, column=1, sticky="ew", padx=5) 
            ctk.CTkLabel(self.search_results_frame, text=mob_version or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i + 1, column=2, sticky="ew", padx=5) 
            ctk.CTkLabel(self.search_results_frame, text=puertas_version or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i + 1, column=3, sticky="ew", padx=5) 
            ctk.CTkLabel(self.search_results_frame, text=fecha_inicio or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i + 1, column=4, sticky="ew", padx=5) 
            ctk.CTkLabel(self.search_results_frame, text=fecha_apertura or 'N/A', font=ctk.CTkFont(weight="bold")).grid(row=i + 1, column=5, sticky="ew", padx=5)

    def create_export_import_frame(self, parent_frame):
        """
        Crea y retorna el frame para la funcionalidad de exportación/importación.
        """
        export_import_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        export_import_frame.grid_columnconfigure(0, weight=1)
        export_import_frame.grid_rowconfigure(3, weight=1) # Espacio para expandir

        ctk.CTkLabel(export_import_frame, text="Exportar / Importar Datos", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, pady=(0, 20))

        # Botones de exportación
        ctk.CTkButton(export_import_frame, text="EXPORTAR A CSV", command=self._export_to_csv, font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, pady=10, padx=20, sticky="ew") # Negrita
        ctk.CTkButton(export_import_frame, text="EXPORTAR A EXCEL", command=self._export_to_excel, font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, pady=10, padx=20, sticky="ew") # Negrita
        
        # Botón de importación (placeholder)
        ctk.CTkButton(export_import_frame, text="IMPORTAR", state="disabled", font=ctk.CTkFont(weight="bold")).grid(row=4, column=0, pady=20, padx=20, sticky="ew") # Negrita
        ctk.CTkLabel(export_import_frame, text="(Funcionalidad de importación se añadirá en el futuro)", font=ctk.CTkFont(size=12, slant="italic", weight="bold")).grid(row=5, column=0, pady=(0,10)) # Negrita

        return export_import_frame

    def _export_to_csv(self):
        """
        Exporta todos los datos de la tabla 'projects' a un archivo CSV.
        """
        if self.current_user_role == 'Visitor':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para exportar datos.")
            logging.warning(f"Intento de exportación a CSV denegado para el rol '{self.current_user_role}'.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*")],
            title="Guardar como archivo CSV"
        )
        if not file_path:
            logging.info("Exportación a CSV cancelada.")
            return # Usuario canceló

        conn = None
        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()

            # Obtener nombres de las columnas de la tabla 'projects'
            cursor.execute("PRAGMA table_info(projects)")
            project_column_names = [col[1] for col in cursor.fetchall()]

            # Obtener nombres de las columnas de la tabla 'project_updates_history'
            cursor.execute("PRAGMA table_info(project_updates_history)")
            history_column_names = [col[1] for col in cursor.fetchall()]

            # Obtener todos los datos de 'projects'
            cursor.execute("SELECT * FROM projects")
            project_rows = cursor.fetchall()

            # Obtener todos los datos de 'project_updates_history'
            cursor.execute("SELECT * FROM project_updates_history")
            history_rows = cursor.fetchall()

            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                csv_writer = csv.writer(csvfile)
                
                # Escribir datos de la tabla 'projects'
                csv_writer.writerow(["--- PROJECTS TABLE ---"])
                csv_writer.writerow(project_column_names) # Escribir encabezados
                csv_writer.writerows(project_rows) # Escribir datos
                
                csv_writer.writerow([]) # Fila vacía para separación
                
                # Escribir datos de la tabla 'project_updates_history'
                csv_writer.writerow(["--- PROJECT UPDATES HISTORY TABLE ---"])
                csv_writer.writerow(history_column_names) # Escribir encabezados
                csv_writer.writerows(history_rows) # Escribir datos

            logging.info(f"Datos exportados a CSV: {file_path} por usuario ID {self.current_user_id}.")
            messagebox.showinfo("Exportación Exitosa", f"Datos exportados a:\n{file_path}")

        except Exception as e:
            logging.error(f"Error al exportar a CSV por usuario ID {self.current_user_id}: {e}")
            messagebox.showerror("Error de Exportación", f"Ocurrió un error al exportar a CSV:\n{e}")
        finally:
            if conn:
                conn.close()

    def _get_summary_data(self):
        """
        Fetches all raw summary data from the database.
        Includes project ID for linking to details.
        """
        conn = None
        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT
                    p.id, -- Added project ID here
                    p.codigo,
                    p.pais,
                    p.ciudad,
                    p.fecha_inicio_obra,
                    p.fecha_apertura,
                    p.image_model_version,
                    p.constructive_model_version,
                    p.mobiliario_version_propuesta,
                    p.mobiliario_actualizacion_date,
                    p.puertas_version_propuesta,
                    p.puertas_actualizacion_date,
                    (SELECT peticionario_name FROM project_updates_history AS puh2 WHERE puh2.project_id = p.id ORDER BY puh2.update_date DESC, puh2.revision_version_at_update DESC LIMIT 1) AS last_peticionario,
                    p.revision_version
                FROM
                    projects AS p
                ORDER BY p.codigo ASC
            """) # Added ORDER BY for default sorting
            summary_data = cursor.fetchall()
            return summary_data
        except Exception as e:
            logging.error(f"Error de base de datos al cargar el resumen de proyectos: {e}")
            messagebox.showerror("Error de Base de Datos", f"Ocurrió un error al cargar el resumen de proyectos:\n{e}")
            return []
        finally:
            if conn:
                conn.close()

    def _sort_summary_table(self, column_key):
        """
        Sorts the summary table by the given column_key.
        """
        if self.current_sort_column_key == column_key:
            self.current_sort_order = "desc" if self.current_sort_order == "asc" else "asc"
        else:
            self.current_sort_column_key = column_key
            self.current_sort_order = "asc" # Default to ascending for new column

        self.load_summary_data() # Reload data with new sort order

    def _export_to_excel(self):
        """
        Exporta los datos tal cual se ven en el resumen a un archivo Excel (.xlsx),
        incluyendo encabezados de dos niveles y la fecha del reporte.
        """
        if self.current_user_role == 'Visitor':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para exportar datos.")
            logging.warning(f"Intento de exportación a Excel denegado para el rol '{self.current_user_role}'.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            title="Guardar Reporte de Resumen BIM como Excel"
        )
        if not file_path:
            logging.info("Exportación a Excel cancelada.")
            return # Usuario canceló

        try:
            raw_summary_data = self._get_summary_data()

            if not raw_summary_data:
                messagebox.showinfo("Información", "No hay proyectos registrados para exportar.")
                return

            # Apply sorting based on current UI state for export
            sort_info = self.sortable_columns_map.get(self.current_sort_column_key)
            if sort_info:
                db_index = sort_info["db_index"]
                sort_type = sort_info["type"]
                reverse = (self.current_sort_order == "desc")

                def get_sort_key(item):
                    actual_db_index = db_index
                    if sort_type == "date":
                        date_str = item[actual_db_index] if not isinstance(actual_db_index, tuple) else item[actual_db_index[0]]
                        try:
                            return datetime.strptime(date_str, '%Y-%m-%d') if date_str and date_str != 'N/A' else datetime.min
                        except ValueError:
                            return datetime.min
                    elif sort_type == "numeric":
                        return item[actual_db_index] if item[actual_db_index] is not None else -float('inf')
                    elif sort_type == "string":
                        if isinstance(actual_db_index, tuple):
                            return " ".join(str(item[idx]) if item[idx] is not None else '' for idx in actual_db_index)
                        return item[actual_db_index] if item[actual_db_index] is not None else ''
                    return item[actual_db_index]

                summary_data_sorted = sorted(raw_summary_data, key=get_sort_key, reverse=reverse)
            else:
                summary_data_sorted = raw_summary_data

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Resumen Proyectos BIM"

            # --- Estilos para Excel ---
            font_bold = Font(bold=True)
            font_bold_underline = Font(bold=True, underline="single")
            fill_header_level1 = PatternFill(start_color="1A4A8C", end_color="1A4A8C", fill_type="solid") # Azul más oscuro
            fill_header_level2 = PatternFill(start_color="2D68C4", end_color="2D68C4", fill_type="solid") # Azul original
            fill_row_even = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Gris claro para filas pares
            fill_row_odd = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Gris un poco más claro para filas impares
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # --- Fecha del Reporte ---
            current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sheet.merge_cells('A1:K1') # Merge cells for the date
            date_cell = sheet['A1']
            date_cell.value = f"Reporte Generado: {current_datetime}"
            date_cell.font = font_bold
            date_cell.alignment = align_center
            
            # Dejar una fila vacía para separación
            sheet.append([]) 

            # --- Encabezados de Nivel Superior (Fila 3 en Excel, después de la fecha y fila vacía) ---
            sheet.merge_cells('A3:A4'); sheet['A3'].value = "PROYECTO"
            sheet.merge_cells('B3:C3'); sheet['B3'].value = "FECHAS DE OBRA"
            sheet.merge_cells('D3:E3'); sheet['D3'].value = "PROYECTO BIM"
            sheet.merge_cells('F3:G3'); sheet['F3'].value = "MOBILIARIO"
            sheet.merge_cells('H3:I3'); sheet['H3'].value = "PUERTAS"
            sheet.merge_cells('J3:K3'); sheet['J3'].value = "ACTUALIZACIÓN"

            for cell in ['A3', 'B3', 'D3', 'F3', 'H3', 'J3']:
                sheet[cell].font = font_bold_underline
                sheet[cell].fill = fill_header_level1
                sheet[cell].alignment = align_center

            # --- Encabezados de Nivel Inferior (Fila 4 en Excel) ---
            headers_row2_display_for_excel = [col[0] for col in self.sortable_column_definitions]
            excel_col_start_for_row2 = 2
            for col_idx_in_list, header_text in enumerate(headers_row2_display_for_excel[1:]):
                current_excel_col = excel_col_start_for_row2 + col_idx_in_list
                cell = sheet.cell(row=4, column=current_excel_col, value=header_text)
                cell.font = font_bold
                cell.fill = fill_header_level2
                cell.alignment = align_center

            # --- Datos de la tabla (a partir de la Fila 5) ---
            start_row_data = 5
            for row_idx, row_data_db in enumerate(summary_data_sorted):
                current_excel_row = start_row_data + row_idx
                current_row_fill = fill_row_even if row_idx % 2 == 0 else fill_row_odd

                project_display_name = f"{row_data_db[1]} {row_data_db[2]} {row_data_db[3]}"
                values_to_write = [project_display_name] + [val or 'N/A' for val in row_data_db[4:]]

                for col_idx, value in enumerate(values_to_write):
                    cell = sheet.cell(row=current_excel_row, column=col_idx + 1, value=str(value))
                    cell.font = font_bold if col_idx == 0 else Font()
                    cell.fill = current_row_fill
                    cell.alignment = align_left if col_idx == 0 else align_center

            # --- Ajustar ancho de columnas ---
            column_widths = {1: 56, 2: 15, 3: 15, 4: 10, 5: 10, 6: 8, 7: 15, 8: 8, 9: 15, 10: 22, 11: 12}
            for i, width in column_widths.items():
                sheet.column_dimensions[get_column_letter(i)].width = width

            workbook.save(file_path)
            logging.info(f"Reporte de resumen exportado a Excel: {file_path} por usuario ID {self.current_user_id}.")
            messagebox.showinfo("Exportación Exitosa", f"Reporte de resumen exportado a:\n{file_path}")

        except Exception as e:
            logging.error(f"Error al exportar a Excel por usuario ID {self.current_user_id}: {e}")
            messagebox.showerror("Error de Exportación", f"Ocurrió un error al exportar a Excel:\n{e}")

    def create_summary_frame(self, parent_frame):
        """
        Crea y retorna el frame para la sección de resumen de proyectos.
        """
        summary_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        summary_frame.grid_columnconfigure(0, weight=1)
        summary_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(summary_frame, text="RESUMEN DE PROYECTOS", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, pady=(10, 15), padx=10)

        # --- Tabla de Resumen ---
        self.summary_table_frame = ctk.CTkScrollableFrame(summary_frame, corner_radius=10, fg_color=("gray80", "gray20"))
        self.summary_table_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        
        self.summary_table_frame.grid_columnconfigure(0, weight=0, minsize=300) 
        self.summary_table_frame.grid_columnconfigure(1, weight=1)
        self.summary_table_frame.grid_columnconfigure(2, weight=1)
        self.summary_table_frame.grid_columnconfigure(3, weight=1)
        self.summary_table_frame.grid_columnconfigure(4, weight=1)
        self.summary_table_frame.grid_columnconfigure(5, weight=1)
        self.summary_table_frame.grid_columnconfigure(6, weight=1)
        self.summary_table_frame.grid_columnconfigure(7, weight=1)
        self.summary_table_frame.grid_columnconfigure(8, weight=1)
        self.summary_table_frame.grid_columnconfigure(9, weight=2)
        self.summary_table_frame.grid_columnconfigure(10, weight=1)

        self.header_buttons = {}
        return summary_frame

    def load_summary_data(self):
        """
        Carga y muestra los datos de resumen en la tabla con encabezados agrupados.
        """
        for widget in self.summary_table_frame.winfo_children():
            widget.destroy()

        raw_summary_data = self._get_summary_data()

        if not raw_summary_data:
            ctk.CTkLabel(self.summary_table_frame, text="No hay proyectos registrados para mostrar el resumen.", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=11, pady=20)
            return

        # Apply sorting
        sort_info = self.sortable_columns_map.get(self.current_sort_column_key)
        if sort_info:
            db_index = sort_info["db_index"] 
            sort_type = sort_info["type"]
            reverse = (self.current_sort_order == "desc")

            def get_sort_key(item):
                if sort_type == "date":
                    date_str = item[db_index] if not isinstance(db_index, tuple) else item[db_index[0]]
                    try:
                        return datetime.strptime(date_str, '%Y-%m-%d') if date_str and date_str != 'N/A' else datetime.min
                    except ValueError:
                        return datetime.min
                elif sort_type == "numeric":
                    return item[db_index] if item[db_index] is not None else -float('inf')
                elif sort_type == "string":
                    if isinstance(db_index, tuple):
                        return " ".join(str(item[idx]) if item[idx] is not None else '' for idx in db_index)
                    return str(item[db_index]) if item[db_index] is not None else ''
                return item[db_index]

            summary_data_sorted = sorted(raw_summary_data, key=get_sort_key, reverse=reverse)
        else:
            summary_data_sorted = raw_summary_data
        
        header_font_level1 = ctk.CTkFont(size=12, weight="bold", underline=True)
        header_font_level2 = ctk.CTkFont(size=12, weight="bold")
        header_bg_color_level1 = "#1A4A8C"
        header_bg_color_level2 = "#2D68C4"
        row_bg_color_even = ("gray75", "gray25")
        row_bg_color_odd = ("gray85", "gray15")

        # --- Encabezados de Nivel Superior ---
        ctk.CTkLabel(self.summary_table_frame, text="PROYECTO", font=header_font_level1, fg_color=header_bg_color_level1, corner_radius=5).grid(row=0, column=0, rowspan=2, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.summary_table_frame, text="FECHAS DE OBRA", font=header_font_level1, fg_color=header_bg_color_level1, corner_radius=5).grid(row=0, column=1, columnspan=2, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.summary_table_frame, text="PROYECTO BIM", font=header_font_level1, fg_color=header_bg_color_level1, corner_radius=5).grid(row=0, column=3, columnspan=2, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.summary_table_frame, text="MOBILIARIO", font=header_font_level1, fg_color=header_bg_color_level1, corner_radius=5).grid(row=0, column=5, columnspan=2, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.summary_table_frame, text="PUERTAS", font=header_font_level1, fg_color=header_bg_color_level1, corner_radius=5).grid(row=0, column=7, columnspan=2, sticky="nsew", padx=1, pady=1)
        ctk.CTkLabel(self.summary_table_frame, text="ACTUALIZACIÓN", font=header_font_level1, fg_color=header_bg_color_level1, corner_radius=5).grid(row=0, column=9, columnspan=2, sticky="nsew", padx=1, pady=1)

        # --- Encabezados de Nivel Inferior ---
        self.header_buttons = {} 
        for col_idx, (base_text, column_key, _) in enumerate(self.sortable_column_definitions):
            if col_idx == 0: continue # Skip first column as it's merged
            display_text = base_text
            if column_key == self.current_sort_column_key:
                arrow = " ▲" if self.current_sort_order == "asc" else " ▼"
                display_text += arrow
            btn = ctk.CTkButton(self.summary_table_frame, text=display_text, font=header_font_level2, fg_color=header_bg_color_level2, corner_radius=5, command=lambda ck=column_key: self._sort_summary_table(ck))
            btn.grid(row=1, column=col_idx, sticky="nsew", padx=1, pady=1)
            self.header_buttons[column_key] = btn

        # --- Datos de la tabla ---
        for row_idx, row_data_db in enumerate(summary_data_sorted):
            current_row_bg_color = row_bg_color_even if row_idx % 2 == 0 else row_bg_color_odd
            project_id = row_data_db[0]
            project_name_display = f"{row_data_db[1]} {row_data_db[2]} {row_data_db[3]}"
            values = [project_name_display] + [val or 'N/A' for val in row_data_db[4:]]

            for col_idx, value in enumerate(values):
                if col_idx == 0:
                    btn = ctk.CTkButton(self.summary_table_frame, text=str(value), command=lambda p_id=project_id: (self.select_frame_by_name("registered_projects"), self.display_project_details(p_id)), fg_color="transparent", text_color_disabled=("gray", "gray"), font=ctk.CTkFont(weight="bold", underline=True), anchor="w")
                    btn.grid(row=row_idx + 2, column=col_idx, sticky="nsew", padx=2, pady=1)
                else:
                    ctk.CTkLabel(self.summary_table_frame, text=str(value), anchor="center", fg_color=current_row_bg_color, font=ctk.CTkFont(weight="bold")).grid(row=row_idx + 2, column=col_idx, sticky="nsew", padx=1, pady=1)


    def create_settings_frame(self, parent_frame):
        """
        Crea y retorna el frame para la sección de Ajustes.
        """
        settings_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        settings_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(settings_frame, text="Ajustes de la Aplicación", font=ctk.CTkFont(size=24, weight="bold")).grid(row=0, column=0, pady=20, padx=20)
        
        # --- Frame de Apariencia ---
        appearance_frame = ctk.CTkFrame(settings_frame)
        appearance_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        appearance_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(appearance_frame, text="Apariencia", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10,5), sticky="w")
        
        ctk.CTkLabel(appearance_frame, text="Modo (Claro/Oscuro):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkOptionMenu(appearance_frame, values=["Light", "Dark", "System"], command=self.change_appearance_mode).grid(row=1, column=1, padx=10, pady=10, sticky="w")

        ctk.CTkLabel(appearance_frame, text="Tema de Color:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        ctk.CTkOptionMenu(appearance_frame, values=["blue", "green", "dark-blue"], command=self.change_theme).grid(row=2, column=1, padx=10, pady=10, sticky="w")

        # --- Frame de Copias de Seguridad ---
        backup_frame = ctk.CTkFrame(settings_frame)
        backup_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        backup_frame.grid_columnconfigure(0, weight=1)
        backup_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(backup_frame, text="Copia de Seguridad y Restauración", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10,5), sticky="w")
        
        btn_create_backup = ctk.CTkButton(backup_frame, text="Crear Copia de Seguridad", command=self._create_backup, font=ctk.CTkFont(weight="bold"))
        btn_create_backup.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        
        btn_restore_backup = ctk.CTkButton(backup_frame, text="Restaurar Copia de Seguridad", command=self._restore_backup, fg_color="orange", font=ctk.CTkFont(weight="bold"))
        btn_restore_backup.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        # Disable backup/restore for non-Admin users
        if self.current_user_role != 'Admin':
            btn_create_backup.configure(state="disabled")
            btn_restore_backup.configure(state="disabled")


        # --- Frame de Registro de Actividad ---
        log_frame = ctk.CTkFrame(settings_frame)
        log_frame.grid(row=3, column=0, padx=20, pady=10, sticky="ew")
        log_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(log_frame, text="Registro de Actividad", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=10, pady=(10,5), sticky="w")
        ctk.CTkButton(log_frame, text="Ver Registro de Actividad", command=self._view_log_file, font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=10, pady=10, sticky="ew")


        return settings_frame

    def change_appearance_mode(self, new_appearance_mode: str):
        """Cambia el modo de apariencia."""
        ctk.set_appearance_mode(new_appearance_mode)
        logging.info(f"Modo de apariencia cambiado a: {new_appearance_mode}")
        # Actualizar gráficos si el dashboard está visible
        if "dashboard" in self.frames and self.frames["dashboard"].winfo_viewable():
            self._update_dashboard_chart()

    def change_theme(self, new_theme: str):
        """
        Cambia el tema de color de la aplicación.
        """
        ctk.set_default_color_theme(new_theme)
        logging.info(f"Tema cambiado a: {new_theme} por usuario ID {self.current_user_id}.")
        messagebox.showinfo("Cambio de Tema", "El tema se ha cambiado. Algunos elementos pueden requerir reiniciar la aplicación para actualizarse completamente.")
    
    def _create_backup(self):
        """Crea una copia de seguridad de la base de datos."""
        if self.current_user_role != 'Admin':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para crear copias de seguridad.")
            logging.warning(f"Intento de creación de copia de seguridad denegado para el rol '{self.current_user_role}'.")
            return

        default_filename = f"bim_projects_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".db",
            filetypes=[("Archivos de Base de Datos", "*.db"), ("Todos los archivos", "*.*")],
            initialfile=default_filename,
            title="Guardar Copia de Seguridad de la Base de Datos"
        )
        if not file_path:
            logging.info("Creación de copia de seguridad cancelada.")
            return

        try:
            shutil.copyfile(DB_NAME, file_path)
            logging.info(f"Copia de seguridad creada exitosamente en: {file_path} por usuario ID {self.current_user_id}.")
            messagebox.showinfo("Copia de Seguridad", f"Copia de seguridad creada exitosamente en:\n{file_path}")
        except Exception as e:
            logging.error(f"Error al crear copia de seguridad por usuario ID {self.current_user_id}: {e}")
            messagebox.showerror("Error", f"Error al crear copia de seguridad:\n{e}")

    def _restore_backup(self):
        """Restaura la base de datos desde un archivo de copia de seguridad."""
        if self.current_user_role != 'Admin':
            messagebox.showwarning("Acceso Denegado", "No tiene permisos para restaurar copias de seguridad.")
            logging.warning(f"Intento de restauración de copia de seguridad denegado para el rol '{self.current_user_role}'.")
            return

        file_path = filedialog.askopenfilename(
            defaultextension=".db",
            filetypes=[("Archivos de Base de Datos", "*.db"), ("Todos los archivos", "*.*")],
            title="Seleccionar Archivo de Copia de Seguridad para Restaurar"
        )
        if not file_path:
            logging.info("Restauración de copia de seguridad cancelada.")
            return

        if messagebox.askyesno("Confirmar Restauración", 
                               "ADVERTENCIA: Esto reemplazará la base de datos actual con la copia de seguridad seleccionada. ¿Desea continuar?"):
            try:
                shutil.copyfile(file_path, DB_NAME)
                logging.info(f"Base de datos restaurada exitosamente desde: {file_path} por usuario ID {self.current_user_id}.")
                messagebox.showinfo("Restauración Exitosa", "Base de datos restaurada exitosamente. La aplicación se reiniciará para aplicar los cambios.")
                self.destroy()
                python = sys.executable
                os.execl(python, python, *sys.argv)
            except Exception as e:
                logging.error(f"Error al restaurar copia de seguridad por usuario ID {self.current_user_id}: {e}")
                messagebox.showerror("Error", f"Error al restaurar copia de seguridad:\n{e}\nAsegúrese de que la aplicación no esté usando la base de datos.")

    def _view_log_file(self):
        """Abre el archivo de registro de actividad."""
        try:
            if sys.platform == "win32":
                os.startfile(log_file)
            elif sys.platform == "darwin": # macOS
                import subprocess
                subprocess.call(["open", log_file])
            else: # linux variants
                import subprocess
                subprocess.call(["xdg-open", log_file])
            logging.info(f"Archivo de log abierto: {log_file} por usuario ID {self.current_user_id}.")
        except Exception as e:
            logging.error(f"Error al abrir el archivo de log por usuario ID {self.current_user_id}: {e}")
            messagebox.showerror("Error", f"No se pudo abrir el archivo de registro:\n{e}")
            
    def create_help_frame(self, parent_frame):
        """
        Crea y retorna el frame para la sección de Ayuda con un diseño de tarjetas mejorado.
        """
        help_frame = ctk.CTkScrollableFrame(parent_frame, corner_radius=10)
        help_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(help_frame, text="Guía de Usuario y Soporte", font=ctk.CTkFont(size=24, weight="bold")).pack(pady=20, padx=20)

        help_topics = [
            ("1. Agregar un Nuevo Proyecto", """Para registrar un nuevo proyecto en la aplicación:
- Navega a la pestaña **"NUEVO"** en el menú lateral izquierdo.
- Rellena todos los campos obligatorios en la sección **"Nomenclatura del Proyecto"**.
- En la sección **"Modelos y Elementos"**, introduce las versiones y fechas correspondientes.
- Marca las casillas **"Fecha Hoy"** para rellenar automáticamente la fecha actual.
- Una vez que todos los campos estén completos, haz clic en **"Guardar Proyecto"**."""),
            ("2. Actualizar Datos de un Proyecto Existente", """Para modificar la información de un proyecto ya registrado:
- Ve a la pestaña **"PROYECTOS"**.
- Haz clic en el nombre del proyecto en la lista para ver sus detalles.
- En el panel de detalles, haz clic en **"Actualizar Datos del Proyecto"**.
- En la nueva ventana, introduce el nombre del **Peticionario** y modifica los datos necesarios.
- La **Versión Revisión** se incrementará automáticamente con cada cambio.
- Haz clic en **"Aceptar"** para guardar."""),
            ("3. Eliminar un Proyecto", """Para borrar un proyecto de la base de datos:
- Navega a la pestaña **"PROYECTOS"**.
- Selecciona el proyecto que deseas eliminar.
- El botón **"Eliminar Proyecto"** (en rojo) se habilitará. Haz clic en él.
- Confirma la eliminación. **Esta acción es irreversible**.
- **Nota**: Solo los usuarios con rol 'Administrador' o 'Editor' pueden eliminar proyectos."""),
            ("4. Buscar Proyectos", """Para encontrar proyectos específicos:
- Haz clic en **"BUSCAR"** en el menú lateral.
- Usa el campo de búsqueda para filtrar por cualquier texto (código, país, ciudad, etc.).
- Haz clic en el nombre de un proyecto en los resultados para ver sus detalles."""),
            ("5. Exportar / Importar Datos", """Para gestionar la exportación e importación de datos:
- Ve a la pestaña **"EXPORT / IMPORT"**.
- **EXPORTAR A CSV/EXCEL**: Guarda los datos en el formato deseado.
- **IMPORTAR**: Funcionalidad en desarrollo.
- **Nota**: Solo los usuarios con rol 'Administrador' o 'Editor' pueden exportar."""),
            ("6. Resumen de Proyectos", """Para ver un resumen consolidado de todos los proyectos:
- Haz clic en **"RESUMEN"** en el menú lateral.
- Se mostrará una tabla con la información clave de cada proyecto.
- Puedes ordenar la tabla haciendo clic en los encabezados de las columnas."""),
            ("7. Gestión de Usuarios (Solo Administradores)", """Esta pestaña, visible solo para 'Administradores', permite:
- **Añadir, Editar y Eliminar** usuarios de la aplicación.
- Asignar roles: **Administrador, Editor o Visitante**.
- **Cambiar de Usuario** para cerrar la sesión actual y volver al login.""")
        ]

        for title, content in help_topics:
            self._create_help_card(help_frame, title, content)

        return help_frame

    def _create_help_card(self, parent_frame, title, content):
        """
        Crea una tarjeta de ayuda con título y contenido, con un borde azul.
        """
        # MEJORA: Crear una tarjeta con borde azul para cada punto de ayuda.
        card_frame = ctk.CTkFrame(parent_frame, corner_radius=10, border_width=2, border_color="#3B8ED0")
        card_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(card_frame, text=title, font=ctk.CTkFont(size=16, weight="bold"),
                     anchor="w").pack(padx=15, pady=(10, 5), fill="x")
        
        ctk.CTkLabel(card_frame, text=content.strip(), font=ctk.CTkFont(size=13),
                     wraplength=parent_frame.winfo_width() - 80, justify="left", anchor="w").pack(padx=15, pady=(0, 10), fill="x")


    def create_about_frame(self, parent_frame):
        """
        Crea y retorna el frame para la sección "Acerca de".
        """
        about_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        about_frame.grid_columnconfigure(0, weight=1)
        about_frame.grid_rowconfigure(0, weight=1) 

        center_content_frame = ctk.CTkFrame(about_frame, fg_color="transparent")
        center_content_frame.grid(row=0, column=0, sticky="nsew")
        center_content_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(center_content_frame, text="BIM Manager", font=ctk.CTkFont(size=40, weight="bold")).pack(pady=(50, 5))
        ctk.CTkLabel(center_content_frame, text="Versión 2.2 | Gestor de Proyectos", font=ctk.CTkFont(size=16, slant="italic")).pack(pady=(0, 20))
        
        description_text = """Esta aplicación ha sido diseñada y desarrollada para optimizar la gestión
y el seguimiento de proyectos BIM.

Su propósito es centralizar la información, estandarizar los flujos de trabajo
y proporcionar una herramienta intuitiva para todo el equipo."""
        ctk.CTkLabel(center_content_frame, text=description_text, justify="center", wraplength=500,
                     font=ctk.CTkFont(size=14)).pack(pady=(0, 30), padx=20)
        
        ctk.CTkLabel(center_content_frame, text="Desarrollado por: Jose Manuel Caamaño González",
                     font=ctk.CTkFont(size=12)).pack(pady=(0, 10))
        
        website_link = ctk.CTkLabel(center_content_frame, text="Contacto",
                                    font=ctk.CTkFont(size=12, underline=True), text_color="#2D68C4", cursor="hand2")
        website_link.pack(pady=(0, 50))
        website_link.bind("<Button-1>", lambda e: webbrowser.open_new("https://www.linkedin.com/in/jmcaamanog/")) # Example link

        return about_frame

    def create_user_management_frame(self, parent_frame):
        """
        Crea y retorna el frame para la sección de Gestión de Usuarios.
        Solo accesible por administradores.
        """
        user_management_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        user_management_frame.grid_columnconfigure(0, weight=1)
        user_management_frame.grid_rowconfigure(2, weight=1) # Espacio para la lista de usuarios

        ctk.CTkLabel(user_management_frame, text="Gestión de Usuarios", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, pady=(0, 15))

        # Frame para botones de acción de usuario
        action_buttons_frame = ctk.CTkFrame(user_management_frame, fg_color="transparent")
        action_buttons_frame.grid(row=1, column=0, pady=10, padx=10, sticky="ew")
        action_buttons_frame.grid_columnconfigure((0,1,2,3), weight=1) # Added a column for the new button

        self.add_user_button = ctk.CTkButton(action_buttons_frame, text="Añadir Usuario", command=self._open_add_user_dialog, font=ctk.CTkFont(weight="bold"))
        self.add_user_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        self.edit_user_button = ctk.CTkButton(action_buttons_frame, text="Editar Usuario", command=self._open_edit_user_dialog, state="disabled", font=ctk.CTkFont(weight="bold"))
        self.edit_user_button.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        self.delete_user_button_um = ctk.CTkButton(action_buttons_frame, text="Eliminar Usuario", command=self._delete_selected_user_um, fg_color="red", state="disabled", font=ctk.CTkFont(weight="bold"))
        self.delete_user_button_um.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Nuevo botón para cambiar de usuario
        self.change_user_button = ctk.CTkButton(action_buttons_frame, text="Cambiar de Usuario", command=self._change_user_session, fg_color="orange", font=ctk.CTkFont(weight="bold"))
        self.change_user_button.grid(row=0, column=3, padx=5, pady=5, sticky="ew")


        # Frame para la lista de usuarios
        self.user_list_frame = ctk.CTkScrollableFrame(user_management_frame, corner_radius=10, fg_color=("gray80", "gray20"))
        self.user_list_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        self.user_list_frame.grid_columnconfigure((0,1,2), weight=1) # Username, Role, Select button

        self.current_selected_user_id = None # Para rastrear el usuario seleccionado en la lista

        self.load_user_list() # Cargar la lista inicial de usuarios

        return user_management_frame

    def load_user_list(self):
        """Carga y muestra la lista de usuarios."""
        for widget in self.user_list_frame.winfo_children():
            widget.destroy()

        users = get_all_users()

        if not users:
            ctk.CTkLabel(self.user_list_frame, text="No hay usuarios registrados.", font=ctk.CTkFont(weight="bold")).pack(pady=20)
            self.edit_user_button.configure(state="disabled")
            self.delete_user_button_um.configure(state="disabled")
            return

        # Encabezados de la tabla de usuarios
        header_font = ctk.CTkFont(size=13, weight="bold", underline=True)
        header_bg_color = "#2D68C4"
        ctk.CTkLabel(self.user_list_frame, text="USERNAME", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=0, sticky="nsew", padx=1, pady=1, ipadx=5, ipady=5)
        ctk.CTkLabel(self.user_list_frame, text="ROL", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=1, sticky="nsew", padx=1, pady=1, ipadx=5, ipady=5)
        ctk.CTkLabel(self.user_list_frame, text="ACCIONES", font=header_font, fg_color=header_bg_color, corner_radius=5).grid(row=0, column=2, sticky="nsew", padx=1, pady=1, ipadx=5, ipady=5)

        for i, (user_id, username, role) in enumerate(users):
            row_bg_color = ("gray75", "gray25") if i % 2 == 0 else ("gray85", "gray15")
            
            ctk.CTkLabel(self.user_list_frame, text=username, font=ctk.CTkFont(weight="bold"), fg_color=row_bg_color).grid(row=i+1, column=0, sticky="nsew", padx=1, pady=1, ipadx=5, ipady=5)
            ctk.CTkLabel(self.user_list_frame, text=role, font=ctk.CTkFont(weight="bold"), fg_color=row_bg_color).grid(row=i+1, column=1, sticky="nsew", padx=1, pady=1, ipadx=5, ipady=5)
            
            select_btn = ctk.CTkButton(self.user_list_frame, text="Seleccionar", command=lambda uid=user_id: self._select_user_for_action(uid), font=ctk.CTkFont(weight="bold"))
            select_btn.grid(row=i+1, column=2, sticky="nsew", padx=1, pady=1, ipadx=5, ipady=5)

        self.edit_user_button.configure(state="disabled")
        self.delete_user_button_um.configure(state="disabled")
        self.current_selected_user_id = None # Reset selection

    def _select_user_for_action(self, user_id):
        """Selecciona un usuario para editar o eliminar."""
        self.current_selected_user_id = user_id
        self.edit_user_button.configure(state="normal")
        self.delete_user_button_um.configure(state="normal")
        logging.info(f"Usuario ID {user_id} seleccionado para acción.")

    def _open_add_user_dialog(self):
        """Abre el diálogo para añadir un nuevo usuario."""
        AddEditUserDialog(self, mode="add", callback=self.load_user_list)

    def _open_edit_user_dialog(self):
        """Abre el diálogo para editar un usuario existente."""
        if self.current_selected_user_id is None:
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un usuario para editar.")
            return

        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, role FROM users WHERE id = ?", (self.current_selected_user_id,))
        user_data = cursor.fetchone()
        conn.close()

        if user_data:
            AddEditUserDialog(self, mode="edit", user_data=user_data, callback=self.load_user_list)
        else:
            messagebox.showerror("Error", "Usuario no encontrado.")
            logging.error(f"Intento de editar usuario ID {self.current_selected_user_id} fallido: usuario no encontrado.")

    def _delete_selected_user_um(self):
        """Elimina el usuario seleccionado de la lista de gestión de usuarios."""
        if self.current_selected_user_id is None:
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un usuario para eliminar.")
            return
        
        if self.current_selected_user_id == self.current_user_id:
            messagebox.showwarning("Error", "No puedes eliminar tu propia cuenta mientras estás logueado.")
            logging.warning(f"Intento de eliminar la propia cuenta (ID: {self.current_user_id}) denegado.")
            return

        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT username FROM users WHERE id = ?", (self.current_selected_user_id,))
        username_to_delete = cursor.fetchone()[0]
        conn.close()

        if messagebox.askyesno("Confirmar Eliminación", f"¿Está seguro de que desea eliminar al usuario '{username_to_delete}'? Esta acción es irreversible."):
            if delete_user_from_db(self.current_selected_user_id):
                messagebox.showinfo("Éxito", f"Usuario '{username_to_delete}' eliminado correctamente.")
                logging.info(f"Usuario '{username_to_delete}' (ID: {self.current_selected_user_id}) eliminado por usuario ID {self.current_user_id}.")
                self.load_user_list()
            else:
                messagebox.showerror("Error", "No se pudo eliminar el usuario.")
                logging.error(f"Fallo al eliminar usuario '{username_to_delete}' (ID: {self.current_selected_user_id}) por usuario ID {self.current_user_id}.")

    def _change_user_session(self):
        """
        Cierra la sesión actual y vuelve a la pantalla de inicio de sesión
        para permitir el cambio de usuario.
        """
        if messagebox.askyesno("Cambiar de Usuario", "¿Desea cerrar la sesión actual y cambiar de usuario?"):
            logging.info(f"Usuario ID {self.current_user_id} solicitó cambiar de sesión.")
            self.destroy() # Destroy the current application instance
            python = sys.executable
            # Re-run the script to show the login screen again
            os.execl(python, python, *sys.argv)
    
    # --- Métodos del Dashboard ---
    def _get_theme_colors(self):
        """Obtiene los colores del tema actual para los gráficos."""
        mode = ctk.get_appearance_mode()
        if mode == "Dark":
            return {
                "bg_color": "#2B2B2B",
                "text_color": "#FFFFFF",
                "spine_color": "#565b5e",
                "grid_color": "#3c3f41",
                "bar_color": "#2D68C4",
                "line_color": "#7FFFD4", # Aquamarine
                "marker_color": "#F0E68C" # Khaki
            }
        else: # Light mode
            return {
                "bg_color": "#EBEBEB",
                "text_color": "#000000",
                "spine_color": "#aab0b5",
                "grid_color": "#d6d6d6",
                "bar_color": "#1F6AA5",
                "line_color": "#d9534f", # Reddish
                "marker_color": "#5cb85c" # Greenish
            }

    def create_dashboard_frame(self, parent_frame):
        """Crea y retorna el frame para el dashboard gráfico."""
        dashboard_frame = ctk.CTkFrame(parent_frame, corner_radius=10)
        dashboard_frame.grid_columnconfigure(0, weight=1)
        dashboard_frame.grid_rowconfigure(1, weight=1)

        # --- Frame para título y controles ---
        controls_frame = ctk.CTkFrame(dashboard_frame)
        controls_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        controls_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(controls_frame, text="Seleccionar Gráfico:", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=10, pady=10)
        
        chart_options = [
            "Proyectos por País", 
            "Versiones de Modelo Imagen", 
            "Versiones de Modelo Constructivo",
            "Versiones de Mobiliario",
            "Versiones de Puertas",
            "Proyectos por Mes de Inicio",
            "Proyectos por Año de Inicio",
            "Proyectos por Peticionario",
            "Versiones de Revisión",
            "Proyectos por Mes de Alta"
        ]
        self.chart_selection_menu = ctk.CTkOptionMenu(controls_frame, 
                                                      values=chart_options,
                                                      command=self._update_dashboard_chart)
        self.chart_selection_menu.grid(row=0, column=1, padx=10, pady=10, sticky="w")

        # --- Frame contenedor para el gráfico ---
        self.chart_container_frame = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
        self.chart_container_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.chart_container_frame.grid_columnconfigure(0, weight=1)
        self.chart_container_frame.grid_rowconfigure(0, weight=1)

        self.current_chart_canvas = None

        return dashboard_frame

    def _update_dashboard_chart(self, chart_name=None):
        """Limpia y dibuja el gráfico seleccionado en el dashboard."""
        if self.current_chart_canvas:
            self.current_chart_canvas.get_tk_widget().destroy()
            plt.close('all')

        if chart_name is None:
            chart_name = self.chart_selection_menu.get()

        logging.info(f"Generando gráfico del dashboard: {chart_name}")

        fig = None
        if chart_name == "Proyectos por País":
            fig = self._create_bar_chart_from_db("pais", "Proyectos por País", "Número de Proyectos")
        elif chart_name == "Versiones de Modelo Imagen":
            fig = self._create_bar_chart_from_db("image_model_version", "Versiones de Modelo Imagen (MI)", "Cantidad de Proyectos", is_numeric=True)
        elif chart_name == "Versiones de Modelo Constructivo":
            fig = self._create_bar_chart_from_db("constructive_model_version", "Versiones de Modelo Constructivo (MC)", "Cantidad de Proyectos", is_numeric=True)
        elif chart_name == "Versiones de Mobiliario":
            fig = self._create_bar_chart_from_db("mobiliario_version_propuesta", "Versiones de Mobiliario", "Cantidad de Proyectos", is_numeric=True)
        elif chart_name == "Versiones de Puertas":
            fig = self._create_bar_chart_from_db("puertas_version_propuesta", "Versiones de Puertas", "Cantidad de Proyectos", is_numeric=True)
        elif chart_name == "Proyectos por Mes de Inicio":
            fig = self._create_date_chart("fecha_inicio_obra", "Proyectos por Mes de Inicio", "%Y-%m")
        elif chart_name == "Proyectos por Año de Inicio":
            fig = self._create_date_chart("fecha_inicio_obra", "Proyectos por Año de Inicio", "%Y")
        elif chart_name == "Proyectos por Peticionario":
            fig = self._create_bar_chart_from_db("peticionario_name", "Proyectos por Peticionario", "Cantidad de Proyectos", table="project_updates_history")
        elif chart_name == "Versiones de Revisión":
            fig = self._create_bar_chart_from_db("revision_version", "Versiones de Revisión por Proyecto", "Versión", use_value_counts=False)
        elif chart_name == "Proyectos por Mes de Alta":
            fig = self._create_date_chart("creation_date", "Proyectos por Mes de Alta", "%Y-%m")


        if fig:
            self.current_chart_canvas = FigureCanvasTkAgg(fig, master=self.chart_container_frame)
            self.current_chart_canvas.draw()
            self.current_chart_canvas.get_tk_widget().pack(side=ctk.TOP, fill=ctk.BOTH, expand=True)
        else:
            # Mostrar un mensaje si no hay datos para el gráfico seleccionado
            for widget in self.chart_container_frame.winfo_children():
                widget.destroy()
            ctk.CTkLabel(self.chart_container_frame, text="No hay datos disponibles para este gráfico.", font=ctk.CTkFont(size=16, slant="italic")).pack(expand=True)
    
    def _create_bar_chart_from_db(self, column_name, title, xlabel, table="projects", is_numeric=False, use_value_counts=True):
        """Función genérica para crear gráficos de barras horizontales."""
        conn = sqlite3.connect(DB_NAME)
        
        # Correctly build the query based on the chart type
        if use_value_counts:
            query = f"SELECT {column_name} FROM {table}"
        else:
            # For charts that plot values against project codes
            query = f"SELECT codigo, {column_name} FROM {table}"
            
        df = pd.read_sql_query(query, conn)
        conn.close()

        # Data cleaning
        df.dropna(subset=[column_name], inplace=True)
        # Convert to string before checking for empty strings to avoid errors with numeric types
        df = df[df[column_name].astype(str).str.strip() != '']
        df = df[df[column_name].astype(str).str.strip() != 'N/A']

        if df.empty:
            return None

        if is_numeric:
            df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
            df.dropna(subset=[column_name], inplace=True)

        if use_value_counts:
            data = df[column_name].value_counts().sort_values(ascending=True)
            y_axis = data.index.astype(str)
            x_axis = data.values
        else:
            # This branch is for charts like "Version de Revisión"
            df = df.sort_values(by=column_name, ascending=True)
            y_axis = df["codigo"] # Use 'codigo' for the y-axis
            x_axis = df[column_name]

        colors = self._get_theme_colors()
        cmap = plt.get_cmap('viridis')
        bar_colors = cmap(np.linspace(0.1, 0.9, len(y_axis)))

        fig, ax = plt.subplots(figsize=(12, 8), facecolor=colors["bg_color"])
        ax.barh(y_axis, x_axis, color=bar_colors)
        
        ax.set_title(title, color=colors["text_color"], fontsize=16, weight='bold')
        ax.set_xlabel(xlabel, color=colors["text_color"])
        ax.set_facecolor(colors["bg_color"])
        ax.tick_params(axis='x', colors=colors["text_color"])
        ax.tick_params(axis='y', colors=colors["text_color"])
        ax.grid(axis='x', color=colors["grid_color"], linestyle='--', linewidth=0.5)
        
        # Ensure integer ticks on the x-axis for counts/versions
        ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))

        for spine in ax.spines.values():
            spine.set_edgecolor(colors["spine_color"])
        
        fig.tight_layout()
        return fig

    def _create_date_chart(self, column_name, title, date_format):
        """Función genérica para crear gráficos de barras por fecha."""
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql_query(f"SELECT {column_name} FROM projects", conn)
        conn.close()

        df.dropna(subset=[column_name], inplace=True)
        df[column_name] = pd.to_datetime(df[column_name], errors='coerce')
        df.dropna(subset=[column_name], inplace=True)

        if df.empty:
            return None

        df['date_group'] = df[column_name].dt.strftime(date_format)
        data = df['date_group'].value_counts().sort_index(ascending=False)

        colors = self._get_theme_colors()
        cmap = plt.get_cmap('plasma')
        bar_colors = cmap(np.linspace(0.1, 0.9, len(data)))

        fig, ax = plt.subplots(figsize=(12, 8), facecolor=colors["bg_color"])
        ax.barh(data.index, data.values, color=bar_colors)

        ax.set_title(title, color=colors["text_color"], fontsize=16, weight='bold')
        ax.set_xlabel("Número de Proyectos", color=colors["text_color"])
        ax.set_facecolor(colors["bg_color"])
        ax.tick_params(axis='x', colors=colors["text_color"])
        ax.tick_params(axis='y', colors=colors["text_color"])
        ax.grid(axis='x', color=colors["grid_color"], linestyle='--', linewidth=0.5)
        
        # Ensure integer ticks on the x-axis
        ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))

        for spine in ax.spines.values():
            spine.set_edgecolor(colors["spine_color"])
        
        fig.tight_layout()
        return fig


class AddEditUserDialog(ctk.CTkToplevel):
    """
    Ventana de diálogo para añadir o editar usuarios.
    """
    def __init__(self, parent, mode, user_data=None, callback=None):
        super().__init__(parent)
        self.mode = mode # "add" or "edit"
        self.user_data = user_data # (id, username, role) for edit mode
        self.callback = callback
        self.grab_set()
        self.resizable(False, False)

        if self.mode == "add":
            self.title("Añadir Nuevo Usuario")
        else:
            self.title(f"Editar Usuario: {user_data[1]}")

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # Username
        ctk.CTkLabel(self, text="Nombre de Usuario:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.username_entry = ctk.CTkEntry(self)
        self.username_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        # Password
        ctk.CTkLabel(self, text="Contraseña:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.password_entry = ctk.CTkEntry(self, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        if self.mode == "edit":
            ctk.CTkLabel(self, text="(Dejar en blanco para no cambiar)", font=ctk.CTkFont(size=10, slant="italic")).grid(row=2, column=1, padx=10, pady=0, sticky="w")

        # Role
        ctk.CTkLabel(self, text="Rol:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.role_optionmenu = ctk.CTkOptionMenu(self, values=["Admin", "Editor", "Visitor"])
        self.role_optionmenu.grid(row=3, column=1, padx=10, pady=5, sticky="ew")

        # Populate for edit mode
        if self.mode == "edit" and self.user_data:
            self.username_entry.insert(0, self.user_data[1])
            self.role_optionmenu.set(self.user_data[2])

        # Buttons
        btn_save = ctk.CTkButton(self, text="Guardar", command=self._save_user)
        btn_save.grid(row=4, column=0, padx=10, pady=20, sticky="e")
        btn_cancel = ctk.CTkButton(self, text="Cancelar", command=self.destroy)
        btn_cancel.grid(row=4, column=1, padx=10, pady=20, sticky="w")

    def _save_user(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        role = self.role_optionmenu.get()

        if not username:
            messagebox.showerror("Error de Entrada", "El nombre de usuario no puede estar vacío.")
            return
        if self.mode == "add" and not password:
            messagebox.showerror("Error de Entrada", "La contraseña no puede estar vacía para un nuevo usuario.")
            return

        success = False
        if self.mode == "add":
            success = add_user_to_db(username, password, role)
        else: # edit mode
            success = update_user_in_db(self.user_data[0], username, password, role)
        
        if success:
            if self.callback:
                self.callback() # Refresh user list
            self.destroy()

# --- Initialization ---
if __name__ == "__main__":
    logging.info("Iniciando aplicación BIM Manager.")
    
    try:
        # Ensure the script can find files in the same directory when run as an executable
        if getattr(sys, 'frozen', False):
            os.chdir(sys._MEIPASS)
            
        setup_database()
        
        root = ctk.CTk()
        root.withdraw() 
        
        # Show Welcome Screen first
        welcome = WelcomeScreen(root)
        welcome.wait_window()

        if welcome.login_requested:
            login = LoginScreen(root)
            login.wait_window() 

            if login.login_successful:
                app = App(login.user_id, login.user_role)
                app.mainloop()
            else:
                if not root.winfo_exists(): # Check if root was destroyed
                    pass
                else:
                    root.destroy()
        else:
            root.destroy()
            
    except Exception as e:
        logging.critical(f"Error fatal al iniciar la aplicación: {e}", exc_info=True)
        messagebox.showerror("Error Crítico", f"No se pudo iniciar la aplicación.\n\nError: {e}\n\nConsulte 'bim_manager_app.log' para más detalles.")
        sys.exit(1)
